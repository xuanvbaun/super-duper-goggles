# -*- coding: utf-8 -*-
"""工艺清单生成程序（可移植版）
用法:
    python run.py "图纸材料表PDF所在文件夹"
    或直接双击 运行.bat 后输入文件夹路径

流程: 扫描源文件夹(主清单PDF + 分/子件PDF) -> OCR -> 表格重建 -> 生成工艺清单Excel
输出: 保存在源文件夹内, 文件名 工艺清单_{设备名称}（{材料表号}）.xlsx
"""
import os, sys, re, json, hashlib, shutil
from collections import defaultdict, Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(SCRIPT_DIR, '工艺清单标准模版-2026.6.30(1).xlsx')
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')

INT_RE = re.compile(r'^(\d+)')
FLOAT_RE = re.compile(r'\d+\.\d+|\d+')
DRAW_RE = re.compile(r'[A-Z0-9]{2,6}-[A-Z0-9]{2,6}-\d{2,4}(?:-\d{2,4})?')
JUNK_RE = re.compile(r'模板文件|第\d+页|^\d{4}/\d+/\d+|26X-R\d+|26X-k\d+|26X-K\d+')

# 表格列边界（PDF 页面统一渲染为 1685x1192）
# 旧模板：数量列在 x≈245；新模板（Q300 销钉焊接图、Q281 变体）：左列左移/右列右移
BOUNDS_A = [
    ('seq', 245), ('qty', 305), ('unit', 340), ('name', 570), ('spec', 950),
    ('std', 1160), ('material', 1315), ('weight', 1415), ('supplier', 1470), ('cat', 99999),
]
BOUNDS_B = [
    ('seq', 210), ('qty', 280), ('unit', 315), ('name', 565), ('spec', 950),
    ('std', 1160), ('material', 1330), ('weight', 1440), ('supplier', 1510), ('cat', 99999),
]

def pick_bounds(lines, h_yc):
    """按表头'数量'列的位置选择模板列边界（旧 x≈245 / 新 x≈215）"""
    for l in lines:
        if abs(l[1] - h_yc) <= 40 and l[6] == '数量':
            return BOUNDS_B if l[0] < 240 else BOUNDS_A
    return BOUNDS_A

def col_of(xc, bounds=BOUNDS_A):
    for cn, b in bounds:
        if xc < b:
            return cn
    return 'cat'

PROFILE_KEYS = ('角钢', '方钢', '槽钢', '工字钢', 'H型钢', '扁钢')

# ============================== 基础工具 ==============================
def parse_int(t):
    m = INT_RE.match(t or '')
    return int(m.group(1)) if m else None

def parse_float(t):
    m = FLOAT_RE.search(t or '')
    return float(m.group(0)) if m else None

def clean_spec(t):
    t = re.sub(r"(?<=\d)'(?=\d)", '1', t)  # OCR 常把数字间的 1 读成撇号（如 L=311'5 → L=3115）
    t = re.sub(r'口(?=\d)|(?<=\d)口', '□', t).replace('?', '')  # 方钢符号□(OCR常读成口)，中文"口"不替换
    t = re.sub(r'(?<=\d):(?=\d)', '.', t)       # 冒号误读小数点（2:8→2.8）
    t = re.sub(r'(?<=\D)Q(?=\d)', 'Ø', t)       # Q 误读 Ø（Q26.9→Ø26.9）
    t = re.sub(r',(?=\s*\d)', ' ', t)           # 逗号后接数字是噪声（DN20, 026.9→DN20 026.9）
    t = re.sub(r'(^|[\s,])(==+|b=|=)(?=\d)', r'\1L=', t)  # =/==/b= 误读 L=（=1205→L=1205）
    t = re.sub(r'<=(?=\d)', 'L=', t)                  # <= 误读 L=（<=200→L=200）
    t = re.sub(r'K=(\d+)(?![°度]|\.\d|\d)', r'L=\1', t)   # K 误读 L（K=40→L=40，保留销轴 K=120°）
    t = re.sub(r'L=(\d+)Q(?=\D|$)', r'L=\g<1>0', t)      # L=173Q→L=1730（Q误读0）
    t = re.sub(r'L=(\d)\.(\d)(\d)', r'L=\1\2\3', t)      # L=8.50→L=850, L=1.070→L=1070
    t = t.replace('ⅡI', 'Ⅱ')
    t = t.replace('L=1.6', 'L=16')
    t = re.sub(r'(?<=\d)X(?=\d)', 'x', t)
    ls = re.findall(r'L=\d+(?:\.\d+)?', t)
    rest = re.sub(r'L=\d+(?:\.\d+)?', '', t)
    rest = re.sub(r'(?<![\d.])0(\d{2,3})(?![\d])', r'Ø\1', rest)
    out = re.sub(r'\s+', ' ', rest).strip()
    return (out + ' ' + ' '.join(ls)).strip()

def clean_name(t):
    t = t.replace('0型', 'O型')
    t = t.replace('层步梯组装图', '二层步梯组装图')   # OCR 漏"二"字
    t = re.sub(r'\(带\s*侧喷口[）)]', '(带侧喷口)', t)
    t = re.sub(r'（带\s*侧喷口[）)]', '(带侧喷口)', t)
    return t.strip()

def clean_std(t):
    t = t.replace('7Q.1', '70.1')
    t = re.sub(r'^[>＞]+', '', t)
    return t.strip()

def normalize_mat(t):
    t = t.replace('（', '(').replace('）', ')')
    return re.sub(r'\s+', '', t)

def normalize_drawing(d):
    return re.sub(r'-[A-Z]$', '', d)

def refs_of(r):
    return re.findall(DRAW_RE, r.get('std', ''))

def line_center(box):
    xs = [p[0] for p in box]; ys = [p[1] for p in box]
    return (min(xs)+max(xs))/2, (min(ys)+max(ys))/2

def line_box(box):
    xs = [p[0] for p in box]; ys = [p[1] for p in box]
    return min(xs), min(ys), max(xs), max(ys)

# ============================== 源文件扫描 ==============================
def drawing_from_filename(f):
    m = re.search(DRAW_RE, f)
    if m:
        return m.group(0)
    return normalize_drawing(re.sub(r'[（(].*?[)）]', '', f).strip())

def scan_sub_pdfs(folder):
    """扫描一个文件夹里的子件PDF，返回 {规范化图号: 路径}"""
    out = {}
    if not os.path.isdir(folder):
        return out
    for f in sorted(os.listdir(folder)):
        if not f.lower().endswith('.pdf'):
            continue
        d = drawing_from_filename(f)
        if d in out and re.search(r'Q281', f):
            continue  # 已有同图号PDF时忽略 Q281 变体（主清单引用 Q280 基础版）
        out[d] = os.path.join(folder, f)
    return out

def find_project(source_dir):
    """返回 (main_pdf, sub_pdfs: dict[规范化图号]->pdf路径)"""
    source_dir = os.path.abspath(source_dir)
    if not os.path.isdir(source_dir):
        print(f'错误: 文件夹不存在 -> {source_dir}')
        sys.exit(1)

    pdfs_top = [f for f in os.listdir(source_dir) if f.lower().endswith('.pdf')]
    sub_pdfs = scan_sub_pdfs(os.path.join(source_dir, '分'))
    # 跨项目回退：在源文件夹的上级目录里查找同图号的子件PDF（如 A300 引用 GDL130-Q300-20）
    # 本项目 分 优先，兄弟项目只做缺失回退（setdefault 不覆盖已存在的同图号）
    parent = os.path.dirname(source_dir)
    if os.path.isdir(parent):
        for d in os.listdir(parent):
            dp = os.path.join(parent, d)
            if not os.path.isdir(dp) or d == os.path.basename(source_dir):
                continue
            if d == '分':  # 同级的分文件夹本身就是另一个项目的子件目录
                ext = scan_sub_pdfs(dp)
            else:          # 兄弟项目文件夹内的 分
                ext = scan_sub_pdfs(os.path.join(dp, '分'))
            for k, v in ext.items():
                sub_pdfs.setdefault(k, v)

    main_pdf = None
    if pdfs_top:
        cand = [f for f in pdfs_top if '总图' in f]
        main_pdf = os.path.join(source_dir, cand[0]) if cand else os.path.join(source_dir, pdfs_top[0])
    if main_pdf is None:
        print('错误: 源文件夹中没有找到 PDF 文件')
        sys.exit(1)
    return main_pdf, sub_pdfs

# ============================== OCR 步骤 ==============================
def do_ocr(pdf_path, cache_dir, name):
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR
    print(f'  [OCR] {os.path.basename(pdf_path)} ...')
    try:
        pdf = pdfium.PdfDocument(pdf_path)
    except Exception:
        print(f'  [跳过] 无法读取PDF（可能被WPS损坏）: {os.path.basename(pdf_path)}')
        return None
    ocr = RapidOCR()
    out = {'file': pdf_path, 'pages': []}
    for i in range(len(pdf)):
        img = pdf[i].render(scale=2.0).to_pil()
        tmp = os.path.join(cache_dir, f'{name}_p{i}.png')
        img.save(tmp)
        res, _ = ocr(tmp)
        page = [{'box': it[0], 'text': it[1], 'conf': float(it[2])} for it in (res or [])]
        out['pages'].append(page)
        print(f'  [OCR]   页{i+1}: {len(page)} 行')
        os.remove(tmp)  # 只保留识别结果，不保留临时图片
    with open(os.path.join(cache_dir, name + '.json'), 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False)
    return out

# ============================== 表格重建 ==============================
def meta_of(doc):
    joined = '\n'.join(it['text'] for it in doc['pages'][0])
    m = {}
    for key, pat in [('drawing', r'材料表号[:：]?\s*([A-Za-z0-9\-]+)'),
                     ('qty', r'制造数量：?\s*([0-9]+)'),
                     ('name', r'设备名称：?\s*([^\n]+)'),
                     ('contract', r'合同号：?\s*([^\n]+)')]:
        mm = re.search(pat, joined)
        if mm: m[key] = mm.group(1).strip()
    return m

def parse_page(page):
    lines = []
    for it in page:
        box, text, conf = it['box'], it['text'], it['conf']
        xc, yc = line_center(box)
        x0, y0, x1, y1 = line_box(box)
        lines.append((xc, yc, y0, y1, x0, x1, text, conf))
    header = next((l for l in lines if l[6] == '名称'), None)
    if header is None:
        header = next((l for l in lines if '名称' in l[6]), None)
    if header is None:
        return None
    h_yc = header[1]
    bounds = pick_bounds(lines, h_yc)
    h_bottom = h_yc + 30
    data = [l for l in lines if l[1] > h_bottom and l[4] > 60 and not JUNK_RE.search(l[6])]
    if not data:
        return []
    data.sort(key=lambda l: l[1])
    bands, cur = [], [data[0]]
    for l in data[1:]:
        if l[1] - cur[-1][1] < 13.5:
            cur.append(l)
        else:
            bands.append(cur); cur = [l]
    bands.append(cur)
    rows = []
    for band in bands:
        rec = {'y': (min(l[1] for l in band) + max(l[1] for l in band)) / 2, 'cols': defaultdict(list)}
        for l in band:
            rec['cols'][col_of(l[0], bounds)].append((l[6], l[7], l[2]))
        rows.append(rec)
    return rows

def build_tables(cache_dir, only_names=None):
    """读取缓存目录里的 OCR 结果重建表格。
    only_names: 只读取这些缓存名（本次实际扫描到的 PDF），防止已删除子件的旧缓存污染结果。"""
    results = {}
    for f in sorted(os.listdir(cache_dir)):
        if not f.endswith('.json'):
            continue
        name = f[:-5]
        if only_names is not None and name not in only_names:
            continue
        with open(os.path.join(cache_dir, f), encoding='utf-8') as fh:
            doc = json.load(fh)
        records = []
        for pi, page in enumerate(doc['pages']):
            rows = parse_page(page)
            if rows is None:
                continue
            for r in rows:
                rec = {}
                for cn in ('seq','qty','unit','name','spec','std','material','weight','supplier','cat'):
                    if cn in r['cols']:
                        vals = sorted(r['cols'][cn], key=lambda v: v[2])
                        text = re.sub(r'\s+', ' ', ' '.join(v[0] for v in vals)).strip()
                        if text:
                            rec[cn] = (text, min(v[1] for v in vals))
                if len(rec) == 1 and 'weight' in rec:
                    continue
                if any('总重' in rec[k][0] for k in rec):
                    continue
                if set(rec.keys()) <= {'supplier', 'cat'}:
                    continue
                rec['_page'] = pi
                rec['_y'] = r['y']
                records.append(rec)
        results[name] = {'file': doc['file'], 'records': records, 'meta': meta_of(doc)}
    return results

# ============================== 重量列二次读取 ==============================
W_RE = re.compile(r'^\d+\.\d{1,3}$')   # 只接受带小数点的干净读数

def _clean_weight(val):
    if W_RE.match(val):
        return float(val)
    return None

def reread_missing_weights(tables):
    """低倍率 OCR 会漏掉部分重量格（小字）。对每页重量列做 scale6 条带 OCR，
    按行 y 对齐（条带 y ≈ 行 y - 16）补全缺失的重量。
    只补'完全没有'的行，且只接受 ^\d+\.\d{1,3}$ 格式，避免噪声写入。"""
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR
    from PIL import Image
    ocr = RapidOCR()
    filled = 0
    for t in tables.values():
        try:
            pdf = pdfium.PdfDocument(t['file'])
        except Exception:
            continue
        by_page = {}
        for rec in t['records']:
            by_page.setdefault(rec.get('_page', 0), []).append(rec)
        for pi, recs in by_page.items():
            img6 = pdf[pi].render(scale=6.0).to_pil()
            k = 3.0
            # x 从 1314 起，避开材质列（如 A16.01.11.A04）右缘渗出；重量值居中于 1320-1405
            strip = img6.crop((int(1314 * k), int(380 * k), int(1412 * k), int(1300 * k)))
            strip = strip.resize((strip.width * 3, strip.height * 3), Image.LANCZOS)
            strip.save('_ws.png')
            try:
                res, _ = ocr('_ws.png')
            finally:
                os.remove('_ws.png')
            readings = []
            if res:
                for it in sorted(res, key=lambda it: it[0][0][1]):
                    v = _clean_weight(it[1].strip())
                    if v is not None:
                        readings.append((380 + it[0][0][1] / 9, v))
            for rec in recs:
                if 'weight' in rec:
                    continue
                y = rec['_y']
                for sy, v in readings:
                    if y - 30 < sy < y - 2:
                        rec['weight'] = (str(v), 1.0)
                        filled += 1
                        break
    if filled:
        print(f'重量列二次读取: 补全 {filled} 个缺失重量')

# ============================== 已确认修正（外置文件） ==============================
def load_corrections():
    p = os.path.join(SCRIPT_DIR, 'rules', 'corrections.json')
    try:
        with open(p, encoding='utf-8') as f:
            return json.load(f).get('corrections', [])
    except Exception as e:
        print(f'[警告] 读取修正文件失败（将跳过已确认修正）: {p} ({e})')
        return []

CORRECTIONS = load_corrections()

def apply_corrections(records, corrections):
    """按 corrections.json 的规则修正记录。
    匹配字段：drawing(图号) / drawings(图号列表) / name / spec_contains / spec_eq / qty_eq / weight_gt / weight_is_null；
    匹配后把 set 里的字段改为新值。"""
    for r in records:
        for c in corrections:
            if c.get('drawing') and c['drawing'] != r['doc_drawing']:
                continue
            if c.get('drawings') and r['doc_drawing'] not in c['drawings']:
                continue
            if c.get('name') and c['name'] != r['name']:
                continue
            if c.get('spec_contains') and not all(s in r['spec'] for s in c['spec_contains']):
                continue
            if c.get('spec_eq') is not None and r['spec'] != c['spec_eq']:
                continue
            if c.get('qty_eq') is not None and r['qty'] != c['qty_eq']:
                continue
            if c.get('weight_gt') is not None and not (r['weight'] is not None and r['weight'] > c['weight_gt']):
                continue
            if c.get('weight_is_null') and r['weight'] is not None:
                continue
            for k, v in c['set'].items():
                r[k] = v

# ============================== 行处理 ==============================
def process_doc(t):
    doc_drawing = normalize_drawing(t['meta'].get('drawing', ''))
    out = []
    for rec in t['records']:
        r = {
            'seq':   parse_int(rec['seq'][0]) if 'seq' in rec else None,
            'qty':   parse_int(rec['qty'][0]) if 'qty' in rec else 1,
            'unit':  rec['unit'][0].replace('祥', '件') if 'unit' in rec else '件',
            'name':  clean_name(rec['name'][0]) if 'name' in rec else '',
            'spec':  clean_spec(rec['spec'][0]) if 'spec' in rec else '',
            'std':   clean_std(rec['std'][0]) if 'std' in rec else '',
            'material': rec['material'][0].strip() if 'material' in rec else '',
            'weight':   parse_float(rec['weight'][0]) if 'weight' in rec else None,
            'supplier': rec['supplier'][0].strip() if 'supplier' in rec else '',
            'cat':      rec['cat'][0].strip() if 'cat' in rec else '',
            'doc_drawing': doc_drawing,
            'ares': False,
        }
        r['ares'] = 'Ares' in r['supplier']
        if not r['name'] and not r['spec'] and not r['std']:
            continue
        if r['qty'] is None:
            r['qty'] = 1
        # 常见 OCR 修正
        r['std'] = r['std'].replace('GDL465-', 'GDL165-')
        if r['name'] == '流体输送用无缝钢管' and '146x4.0' in r['spec']:
            r['spec'] = r['spec'].replace('146x4.0', 'Ø146x4.0')   # 缺Ø(复核确认)
        if r['name'] == '耐热钢板':
            if 'Q154' in r['spec']:
                r['spec'] = '2×Ø154'          # Q为0误读(复核确认)
            elif '2×154' in r['spec']:
                r['spec'] = '2×Ø154'          # 缺Ø(复核确认)
        if r['name'] == '时效炉风机型线2中间风箱' and r['qty'] == 9:
            r['qty'] = 6                 # OCR 把6误读为9(高倍率复核确认, 与子件图纸制造数量一致)
        if r['name'] == '陶瓷纤维方编绳':
            r['material'] = ''           # 材质格在PDF中为空白
        if r['name'] == '耐热钢板' and '2xØ90' in r['spec']:
            r['qty'] = 1                 # 数量格在PDF中为空白
        r['material'] = re.sub(r'\s*26X\S*', '', r['material']).strip()   # 合同号渗出噪声(复核确认)
        r['spec'] = re.sub(r'^\d+\s+(?=90E\(L\))', '', r['spec'])          # 90°弯头前的孤立数字噪声
        out.append(r)
    apply_corrections(out, CORRECTIONS)   # 已人工确认的修正（rules/corrections.json）
    assign_part_no(out)
    return out

def assign_part_no(records):
    n = len(records)
    seqs = [r['seq'] if r['seq'] is not None and 1 <= r['seq'] <= n + 3 else None for r in records]
    counts = Counter(s for s in seqs if s is not None)
    if any(v > 1 for v in counts.values()):
        for i, r in enumerate(records, 1):
            r['part_no'] = i
    else:
        prev = 0
        for r, s in zip(records, seqs):
            if s is not None:
                r['part_no'] = s; prev = s
            else:
                r['part_no'] = prev + 1; prev = r['part_no']

STD_NAME_RE = re.compile(r'螺栓|螺钉|螺母|垫圈|开口销|销轴|铆钉|挡圈|键|轴承|油杯')

def is_std_part(r):
    if refs_of(r):
        return False  # 有子图号引用的是自制件，不是标准件（分类号列常被OCR误读成16）
    if r['cat'] in ('16', '17', '7') and '编绳' not in r['name']:
        return True
    # 分类号列常被 OCR 读坏（如 116 2 4 / 空），用"标准件名称+GB/T/JB/T标准号"兜底
    return (bool(STD_NAME_RE.search(r['name'])) and bool(re.search(r'GB/?T|JB/?T', r['std']))
            and '编绳' not in r['name'])

def classify_material(r):
    name = r['name']
    if '钢管' in name or '焊管' in name or '无缝管' in name:
        return 'pipe'
    if '圆钢' in name:
        return 'bar'
    if '钢板' in name:
        return 'plate'
    if any(k in name for k in PROFILE_KEYS):
        return 'profile'
    if re.match(r'^Ø\d+(?:\.\d+)?x\d+(?:\.\d+)?', r['spec']):
        return 'pipe'
    return None

def plate_thickness(spec):
    nums = [float(m) for m in re.findall(r'\d+(?:\.\d+)?', spec)]
    if not nums:
        return None
    return f"t{min(nums):g}"

def spec_no_length(spec):
    return re.sub(r'\s*L=\d+(?:\.\d+)?', '', spec).strip()

def material_of(r):
    if r['ares']:
        return ''
    cls = classify_material(r)
    if r['material']:
        if cls == 'plate':
            t = plate_thickness(r['spec'])
            return f"钢板{t}/{r['material']}" if t else r['material']
        if cls == 'pipe':
            spec = r['spec']
            if spec and not spec.startswith('Ø') and re.match(r'\d+x', spec):
                spec = 'Ø' + spec
            return f"圆管{spec_no_length(spec)}/{r['material']}"
        if cls == 'bar':
            return f"圆棒{spec_no_length(r['spec'])}/{r['material']}"
        if cls == 'profile':
            return f"{r.get('name_base', r['name'])}{spec_no_length(r['spec'])}/{r['material']}"
    if is_std_part(r):
        parts = []
        if r['spec']:
            parts.append(r['spec'])
        if r['std']:
            s = re.sub(r'图[:：]?|（|）|\(|\)', ' ', r['std'])
            s = re.sub(r'\s+', ' ', s).strip()
            if s:
                parts.append(s)
        if r['material']:
            parts.append(normalize_mat(r['material']))
        return '/'.join(parts)
    return r['material']

# ============================== 高倍率复核 ==============================
VERIFY_COLS_A = {'qty': (240, 310), 'weight': (1310, 1410), 'spec': (560, 960)}
VERIFY_COLS_B = {'qty': (210, 290), 'weight': (1330, 1440), 'spec': (565, 950)}

def norm_cell(cname, t):
    t = (t or '').strip()
    if cname in ('qty', 'weight'):
        m = re.search(r'\d+(?:\.\d+)?', t)
        return m.group(0) if m else ''
    # 规格: 忽略顺序/空格差异, 保留内容差异 —— 比较"数字序列 + 字母符号集合"
    t = t.lower().replace('×', 'x')
    nums = 'N' + '|'.join(sorted(re.findall(r'\d+(?:\.\d+)?', t)))
    chars = 'C' + ''.join(sorted(re.findall(r'[a-zØ]', t)))
    return nums + chars

def verify_cells(pdf_path, doc, cache_dir):
    """对每页的 数量/重量/规格 列做高倍率(scale6)复核。
    返回 (差异列表, 复核单元格总数)"""
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR
    from PIL import Image
    ocr = RapidOCR()
    try:
        pdf = pdfium.PdfDocument(pdf_path)
    except Exception:
        print(f'  [复核跳过] 无法读取PDF(可能被WPS损坏): {os.path.basename(pdf_path)}')
        return [], 0
    diffs, total = [], 0
    for pi, page in enumerate(doc['pages']):
        lines = []
        for it in page:
            box, text, conf = it['box'], it['text'], it['conf']
            xs = [p[0] for p in box]; ys = [p[1] for p in box]
            lines.append(((min(xs)+max(xs))/2, (min(ys)+max(ys))/2, min(ys), max(ys),
                          min(xs), max(xs), text, conf))
        header = next((l for l in lines if l[6] == '名称'), None)
        if header is None:
            header = next((l for l in lines if '名称' in l[6]), None)
        if header is None:
            continue
        h_yc = header[1]
        bounds = pick_bounds(lines, h_yc)
        verify_cols = VERIFY_COLS_B if bounds is BOUNDS_B else VERIFY_COLS_A
        h_bottom = h_yc + 30
        data = [l for l in lines if l[1] > h_bottom and l[4] > 60 and not JUNK_RE.search(l[6])]
        data.sort(key=lambda l: l[1])
        bands, cur = [], [data[0]] if data else []
        for l in data[1:]:
            if l[1] - cur[-1][1] < 13.5:
                cur.append(l)
            else:
                bands.append(cur); cur = [l]
        if cur:
            bands.append(cur)
        img6 = pdf[pi].render(scale=6.0).to_pil()
        for band in bands:
            y0 = min(l[2] for l in band) - 6
            y1 = max(l[3] for l in band) + 6
            cells = {}
            for l in band:
                cells.setdefault(col_of(l[0], bounds), []).append(l)
            name = ''
            if 'name' in cells:
                name = ' '.join(l[6] for l in sorted(cells['name'], key=lambda l: l[2]))
            for cname, (x0, x1) in verify_cols.items():
                if cname not in cells:
                    continue
                vals = sorted(cells[cname], key=lambda l: l[2])
                orig = ' '.join(l[6] for l in vals)
                total += 1
                crop = img6.crop((int(x0 * 3), int(y0 * 3), int(x1 * 3), int(y1 * 3)))
                crop = crop.resize((crop.width * 3, crop.height * 3), Image.LANCZOS)
                crop.save('_v.png')
                res, _ = ocr('_v.png')
                high = ''.join(it[1] for it in (res or []))
                if norm_cell(cname, orig) != norm_cell(cname, high):
                    diffs.append({'page': pi + 1, 'name': name, 'col': cname,
                                  'orig': orig.strip(), 'high': high.strip()})
    if os.path.exists('_v.png'):
        os.remove('_v.png')
    return diffs, total

def run_verify(doc_name, pdf_path, doc, cache_dir, report):
    """执行复核并写入报告列表, 返回 (差异数, 复核格数)"""
    diffs, total = verify_cells(pdf_path, doc, cache_dir)
    if diffs:
        report.append(f'--- {os.path.basename(pdf_path)} ({len(diffs)} 处差异 / 复核 {total} 格) ---')
        for d in diffs:
            report.append(f'  第{d["page"]}页 | {d["name"][:22]} | {d["col"]}: 原识别={d["orig"]}  高倍率={d["high"]}')
    return diffs, total

# ============================== 层级与生成 ==============================
class Node:
    __slots__ = ('row', 'children', 'level', 'mult', 'parent', 'number')
    def __init__(self, row=None):
        self.row = row
        self.children = []
        self.level = 0
        self.mult = 1
        self.parent = None
        self.number = ''

def build_tree(main_records, sub_map, root_draw=''):
    root = Node()
    # 当前递归链上的图号集合（防循环引用 A→B→C→A 无限递归；同一子图被不同父件
    # 重复引用是合法的，所以不能用全局 visited，只用当前路径）
    active = {root_draw} if root_draw else set()

    def attach(node, records):
        for r in records:
            child = Node(r)
            child.parent = node
            child.level = node.level + 1
            child.mult = node.mult * (node.row['qty'] if node.row else 1)
            node.children.append(child)
            for ref in refs_of(r):
                if ref in sub_map:
                    if ref in active:
                        print(f'  [警告] 检测到循环引用: {ref}（本分支停止展开）')
                        break
                    active.add(ref)
                    attach(child, sub_map[ref])
                    active.discard(ref)
                    break
    attach(root, main_records)
    return root

def node_code(node):
    r = node.row
    own = r['doc_drawing']
    for ref in refs_of(r):
        if ref != own:
            return ref
    return f"{own}-{r['part_no']}"

def weight_per(r):
    if r['weight'] is None:
        return None
    return round(r['weight'] / (r['qty'] or 1), 4)

# ============================== 缓存键 ==============================
def file_key(pdf_path):
    """按 规范化绝对路径 + 大小 + 修改时间 生成缓存键。
    同名 PDF（兄弟项目）不撞名；PDF 更新过（size/mtime 变化）自动换新键重新 OCR。"""
    st = os.stat(pdf_path)
    ident = f"{os.path.normcase(os.path.abspath(pdf_path))}|{st.st_size}|{int(st.st_mtime)}"
    return hashlib.md5(ident.encode('utf-8')).hexdigest()[:12]

def run(source_dir):
    main_pdf, sub_pdfs = find_project(source_dir)

    # ---- 格式分流：文字型PDF(如瑞士Rowa BOM)走文本提取+翻译路径 ----
    from pdfminer.high_level import extract_text
    try:
        txt = extract_text(main_pdf) or ''
    except Exception:
        txt = ''
    if len(txt.strip()) > 200 and not os.path.isdir(os.path.join(source_dir, '分')):
        print('检测到文字型 PDF（Rowa/瑞士格式），走文本提取+翻译路径 ...')
        import swiss_build
        out = swiss_build.build_folder(source_dir)
        if out:
            print(f'\n完成! 输出: {out}')
        return

    print(f'主清单: {os.path.basename(main_pdf)}')
    print(f'子件PDF: {len(sub_pdfs)} 个')

    # 缓存目录（按源文件夹路径哈希，重复运行可跳过OCR）
    key = hashlib.md5(source_dir.encode('utf-8')).hexdigest()[:10]
    cache_dir = os.path.join(DATA_DIR, key)
    os.makedirs(cache_dir, exist_ok=True)

    main_name = 'MAIN_' + file_key(main_pdf)
    main_doc = None
    main_json = os.path.join(cache_dir, main_name + '.json')
    if os.path.exists(main_json):
        with open(main_json, encoding='utf-8') as f:
            main_doc = json.load(f)
        print('(使用已缓存的OCR结果)')
    else:
        main_doc = do_ocr(main_pdf, cache_dir, main_name)
    if main_doc is None:
        print(f'\n错误: 主清单PDF无法读取（可能被WPS损坏）: {os.path.basename(main_pdf)}')
        print('请重新获取这份PDF（不要用WPS打开保存它）后再运行。')
        sys.exit(1)
    docs = {main_name: main_doc}
    for d, p in sub_pdfs.items():
        sub_name = 'SUB_' + file_key(p)
        sj = os.path.join(cache_dir, sub_name + '.json')
        if os.path.exists(sj):
            with open(sj, encoding='utf-8') as f:
                docs[sub_name] = json.load(f)
            print(f'(使用已缓存的OCR结果: {os.path.basename(p)})')
        else:
            sub_doc = do_ocr(p, cache_dir, sub_name)
            if sub_doc is None:
                continue  # 子件PDF损坏则跳过（对应图号不展开）
            docs[sub_name] = sub_doc

    # 只读取本次实际扫描到的缓存（防止已删除子件的旧缓存污染结果）
    tables = build_tables(cache_dir, only_names=set(docs.keys()))
    main_t = tables[main_name]

    # ---- 重量列二次读取：补全低倍率漏检的重量格（小字/细笔画）----
    reread_missing_weights(tables)

    # ---- 高倍率复核：数量/重量/规格 列（图片型OCR路径自动执行）----
    print('\n正在做高倍率复核（数量/重量/规格列）...')
    report = ['高倍率复核报告（数量/重量/规格列）', '=' * 60]
    n_diff = n_total = 0
    for sub_name, t in tables.items():
        doc = docs.get(sub_name)
        if doc is None:
            continue
        d, tot = run_verify(sub_name, doc['file'], doc, cache_dir, report)
        n_diff += len(d)
        n_total += tot
    print(f'复核完成: 共 {n_total} 格, 发现 {n_diff} 处与原识别不一致')
    if report:
        rp = os.path.join(source_dir, '复核报告.txt')
        with open(rp, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report))
        print(f'复核报告已保存: {rp}')

    sub_map = {}
    for sub_name, t in tables.items():
        if sub_name.startswith('SUB_'):
            sub_map[normalize_drawing(t['meta'].get('drawing', ''))] = process_doc(t)

    main_records = process_doc(main_t)
    root_draw = normalize_drawing(main_t['meta'].get('drawing', ''))
    root = build_tree(main_records, sub_map, root_draw)

    # 层级序号
    def assign_numbers(parent):
        i = 1
        for c in parent.children:
            c.number = (parent.number + '.' if parent.number else '') + str(i)
            assign_numbers(c)
            i += 1
    assign_numbers(root)

    flat = []
    def flatten(node):
        if node.row:
            flat.append(node)
        for c in node.children:
            flatten(c)
    flatten(root)

    # 同父件重名加序号（标准件除外）
    children_of = defaultdict(list)
    for node in flat:
        children_of[id(node.parent)].append(node)
    for nodes in children_of.values():
        counts = Counter(n.row['name'] for n in nodes)
        seen = defaultdict(int)
        for node in nodes:
            nm = node.row['name']
            node.row['name_base'] = nm
            if counts[nm] > 1 and not is_std_part(node.row):
                seen[nm] += 1
                node.row['name'] = f"{nm}{seen[nm]}"

    root_name = main_t['meta'].get('name', '')
    root_contract = main_t['meta'].get('contract', '')

    def parent_ref(node):
        p = node.parent
        if p is None or p.row is None:
            return (root_draw, root_name)
        return (node_code(p), p.row['name'])

    def remark_of(r):
        return '' if r['ares'] else r['spec']

    # 生成 Excel
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb['模板表单']
    ws['C1'] = f"{root_contract} {root_name}（{root_draw}）"

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    FONT = Font(name='宋体', size=10)
    FONT_B = Font(name='宋体', size=10, bold=True)
    ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

    r0 = 3
    row_of = {id(node): r0 + i for i, node in enumerate(flat)}
    for i, node in enumerate(flat):
        r = node.row
        xr = r0 + i
        pref, pname = parent_ref(node)
        data = {
            1: node.number,
            2: node_code(node),
            3: r['name'],
            4: pref,
            5: pname,
            6: r['qty'],
            7: r['qty'] * node.mult,
            9: material_of(r),
            10: weight_per(r),
            14: remark_of(r),
        }
        bold = bool(node.children)
        for ci in range(1, 20):
            c = ws.cell(xr, ci)
            c.font = FONT_B if bold else FONT
            c.border = border
            c.alignment = ALIGN
            if ci in data:
                c.value = data[ci]
                if ci in (6, 7):
                    c.number_format = '0'
                elif ci == 10:
                    c.number_format = '0.0000'
        # K 列（单台总重）：父件 = 直接子件 K 之和；叶子 = 单台数量×单件重量
        if node.children:
            ws.cell(xr, 11).value = '=SUM(' + ','.join(f'K{row_of[id(c)]}' for c in node.children) + ')'
        else:
            ws.cell(xr, 11).value = f'=G{xr}*J{xr}'
    for xr in range(3, ws.max_row + 1):
        for ci in range(1, 20):
            c = ws.cell(xr, ci)
            if c.fill and c.fill.patternType:
                c.fill = PatternFill()

    out_name = f"工艺清单_{root_name}（{root_draw}）.xlsx"
    out_path = os.path.join(source_dir, out_name)
    try:
        wb.save(out_path)
        print(f'\n完成! 共 {len(flat)} 行')
        print(f'输出: {out_path}')
    except PermissionError:
        alt = os.path.join(source_dir, out_name[:-5] + '（新）.xlsx')
        wb.save(alt)
        print(f'\n注意: 目标文件正被占用(可能WPS打开)，已另存为:\n  {alt}')
        print(f'完成! 共 {len(flat)} 行')

    # 打印前几行供核对
    for node in flat[:8]:
        r = node.row
        print(f"  {node.number:<6}{r['name']:<20} 数量{node.mult * r['qty']}  材质:{material_of(r)}")

if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else input('请输入图纸材料表PDF所在文件夹路径：').strip().strip('"')
    if not src:
        print('未输入路径')
        sys.exit(1)
    run(src)
