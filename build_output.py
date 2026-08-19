# -*- coding: utf-8 -*-
"""Build 工艺清单 Excel v2 (per user's 7 revision rules):
1. 标准件材质 = 规格/标准号/强度/表面处理（缺项跳过）
2. 计划数量列不填
3. 不加颜色
4. 备注 = 规格/设备PI号列
5. Ares(买方供货)行不填（材质保持原值，备注留空）
6. 代号 = 材料表号+件号(去字母)；行内有其他图号则用该图号
7. 同一父件下重名子件，名称后加序号
"""
import json, re, os, sys
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

TARGET = sys.argv[1] if len(sys.argv) > 1 else 'Q300'
BASE = os.path.dirname(os.path.abspath(__file__))
CONF = {
    'Q300': {
        'tables': os.path.join(BASE, 'tables.json'),
        'out_dir': r'E:\李鑫宇\2\新建文件夹',
        'out_file': '工艺清单_固溶炉炉体钢结构（GDL165-Q300-01）.xlsx',
    },
    'A300': {
        'tables': os.path.join(BASE, 'tables_a300.json'),
        'out_dir': r'E:\李鑫宇\2\新建文件夹\阿瑞斯A300',
        'out_file': '工艺清单_时效炉炉体钢结构（GDL165-A300-00）.xlsx',
    },
}[TARGET]

TABLES = json.load(open(CONF['tables'], encoding='utf-8'))
TEMPLATE = r'E:\李鑫宇\2\工艺清单标准模版-2026.6.30(1).xlsx'
OUT_DIR = CONF['out_dir']
OUT_FILE = CONF['out_file']

INT_RE = re.compile(r'^(\d+)')
FLOAT_RE = re.compile(r'\d+\.\d+|\d+')
DRAW_RE = re.compile(r'[A-Z0-9]{2,6}-[A-Z0-9]{2,6}-\d{2,4}(?:-\d{2,4})?')

def parse_int(t):
    m = INT_RE.match(t)
    return int(m.group(1)) if m else None

def parse_float(t):
    m = FLOAT_RE.search(t)
    return float(m.group(0)) if m else None

def clean_spec(t):
    t = t.replace('口', '□').replace('?', '')
    t = t.replace('ⅡI', 'Ⅱ')
    t = t.replace('L=1.6', 'L=16')  # OCR: M6 screw L=1.6 -> L=16
    t = re.sub(r'(?<=\d)X(?=\d)', 'x', t)  # 2X090 -> 2x090
    ls = re.findall(r'L=\d+(?:\.\d+)?', t)
    rest = re.sub(r'L=\d+(?:\.\d+)?', '', t)
    rest = re.sub(r'(?<![\d.])0(\d{2,3})(?![\d])', r'Ø\1', rest)
    out = re.sub(r'\s+', ' ', rest).strip()
    return (out + ' ' + ' '.join(ls)).strip()

def clean_name(t):
    t = t.replace('0型', 'O型')
    t = re.sub(r'\(带\s*侧喷口[）)]', '(带侧喷口)', t)
    t = re.sub(r'（带\s*侧喷口[）)]', '(带侧喷口)', t)
    return t.strip()

def clean_std(t):
    t = t.replace('7Q.1', '70.1')
    t = re.sub(r'^[>＞]+', '', t)
    return t.strip()

def normalize_mat(t):
    t = t.replace('（', '(').replace('）', ')')
    return re.sub(r'\s+', '', t)

def is_junk(rec):
    keys = set(rec.keys())
    if any('总重' in rec[k][0] for k in rec):
        return True
    if keys <= {'supplier', 'cat'}:
        return True
    if keys == {'weight'}:
        return True
    return False

def normalize_drawing(d):
    return re.sub(r'-[A-Z]$', '', d)

def refs_of(r):
    return re.findall(DRAW_RE, r['std'])

def process_doc(t):
    doc_drawing = normalize_drawing(t['meta'].get('drawing', ''))
    out = []
    for rec in t['records']:
        if is_junk(rec):
            continue
        r = {
            'seq':   parse_int(rec['seq'][0]) if 'seq' in rec else None,
            'qty':   parse_int(rec['qty'][0]) if 'qty' in rec else 1,
            'unit':  rec['unit'][0].replace('祥', '件') if 'unit' in rec else '件',
            'name':  clean_name(rec['name'][0]) if 'name' in rec else '',
            'spec':  clean_spec(rec['spec'][0]) if 'spec' in rec else '',
            'std':   clean_std(rec['std'][0]) if 'std' in rec else '',
            'material': rec['material'][0].strip() if 'material' in rec else '',
            'weight':   parse_float(rec['weight'][0]) if 'weight' in rec else None,
            'supplier': rec['supplier'][0].strip() if 'supplier' in rec else '',
            'cat':      rec['cat'][0].strip() if 'cat' in rec else '',
            'conf': min(rec[k][1] for k in rec),
            'doc_drawing': doc_drawing,
            'ares': False,
        }
        r['ares'] = 'Ares' in r['supplier']
        if not r['name'] and not r['spec'] and not r['std']:
            continue
        if r['qty'] is None:
            r['qty'] = 1
        out.append(r)
    assign_part_no(out)
    return out

def assign_part_no(records):
    """件号: 用清单序号（去重/越界修正），缺失时插值；序号重复则退回位置号"""
    n = len(records)
    seqs = [r['seq'] if r['seq'] is not None and 1 <= r['seq'] <= n + 3 else None for r in records]
    counts = Counter(s for s in seqs if s is not None)
    if any(v > 1 for v in counts.values()):
        for i, r in enumerate(records, 1):
            r['part_no'] = i
    else:
        prev = 0
        for r, s in zip(records, seqs):
            if s is not None:
                r['part_no'] = s
                prev = s
            else:
                r['part_no'] = prev + 1
                prev = r['part_no']

MAIN = process_doc(TABLES['MAIN'])
SUB_MAP = {}
for docname, t in TABLES.items():
    if docname.startswith('SUB_'):
        d = t['meta'].get('drawing', '')
        SUB_MAP[normalize_drawing(d)] = process_doc(t)

# root info from main sheet (no hardcoding)
ROOT_DRAW = normalize_drawing(TABLES['MAIN']['meta'].get('drawing', ''))
ROOT_NAME = TABLES['MAIN']['meta'].get('name', '')
ROOT_CONTRACT = TABLES['MAIN']['meta'].get('contract', '')

# cross-project fallback: rows may reference drawings from the other project
# (e.g. A300 炉体光栅装配 -> 图：GDL130-Q300-20, which has a PDF in Q300's 分 folder)
if TARGET == 'A300':
    q300_tables = json.load(open(os.path.join(BASE, 'tables.json'), encoding='utf-8'))
    for docname, t in q300_tables.items():
        if docname.startswith('SUB_'):
            d = normalize_drawing(t['meta'].get('drawing', ''))
            if d not in SUB_MAP:
                SUB_MAP[d] = process_doc(t)

# manual overrides (verified by high-scale re-OCR when applicable)
if TARGET == 'A300':
    # 导风板子件表: 中间风箱数量 OCR 把6误读为9(高倍率复核确认为6, 与其中间风箱图纸制造数量6一致)
    for recs in SUB_MAP.values():
        for r in recs:
            if r['name'] == '时效炉风机型线2中间风箱' and r['qty'] == 9:
                r['qty'] = 6
    # 中间风箱 耐热钢板: 规格 2xQ154 → 2×Ø154 (Q为0误读, 高倍率复核确认)
    for recs in SUB_MAP.values():
        for r in recs:
            if r['name'] == '耐热钢板' and 'Q154' in r['spec']:
                r['spec'] = '2×Ø154'
if TARGET == 'Q300':
    # 总清单 流体输送用无缝钢管: 146x4.0 → Ø146x4.0 (缺Ø, 高倍率复核确认 0146)
    for r in MAIN:
        if r['name'] == '流体输送用无缝钢管' and '146x4.0' in r['spec']:
            r['spec'] = r['spec'].replace('146x4.0', 'Ø146x4.0')
    # 中间风箱 耐热钢板: 2×154 → 2×Ø154 (缺Ø, 高倍率复核确认 2x0154)
    for recs in SUB_MAP.values():
        for r in recs:
            if r['name'] == '耐热钢板' and '2×154' in r['spec']:
                r['spec'] = '2×Ø154'


for r in MAIN:
    r['std'] = r['std'].replace('GDL465-A300', 'GDL165-A300')  # OCR: 1 misread as 4
    if r['name'] == '陶瓷纤维方编绳':
        r['material'] = ''            # material cell is blank in PDF
    if r['name'] == '耐热钢板' and '2xØ90' in r['spec']:
        r['qty'] = 1                  # qty cell is blank in PDF -> 1

# ---------- tree ----------
class Node:
    __slots__ = ('row', 'children', 'level', 'mult', 'parent', 'number', 'sub_ref')
    def __init__(self, row=None):
        self.row = row
        self.children = []
        self.level = 0
        self.mult = 1
        self.parent = None
        self.number = ''
        self.sub_ref = None

def attach_children(node, records):
    for r in records:
        child = Node(r)
        child.parent = node
        child.level = node.level + 1
        child.mult = node.mult * (node.row['qty'] if node.row else 1)
        node.children.append(child)
        for ref in refs_of(r):
            if ref in SUB_MAP:
                child.sub_ref = ref
                attach_children(child, SUB_MAP[ref])
                break

root = Node()
attach_children(root, MAIN)

flat = []
def flatten(node):
    if node.row:
        flat.append(node)
    for c in node.children:
        flatten(c)
flatten(root)

def assign_numbers(parent):
    i = 1
    for c in parent.children:
        c.number = (parent.number + '.' if parent.number else '') + str(i)
        assign_numbers(c)
        i += 1
assign_numbers(root)

def is_std_part(r):
    return r['cat'] in ('16', '17', '7') and '编绳' not in r['name']

# rule 7: same name within same parent -> append index (标准件不加序号, keep base name for 型材 rule)
children_of = defaultdict(list)
for node in flat:
    children_of[id(node.parent)].append(node)
for nodes in children_of.values():
    counts = Counter(node.row['name'] for node in nodes)
    seen = defaultdict(int)
    for node in nodes:
        nm = node.row['name']
        node.row['name_base'] = nm
        if counts[nm] > 1 and not is_std_part(node.row):
            seen[nm] += 1
            node.row['name'] = f"{nm}{seen[nm]}"

# rule 6: 代号 = 材料表号-件号(去字母)，行内有其他图号则用图号
def node_code(node):
    r = node.row
    own = r['doc_drawing']
    for ref in refs_of(r):
        if ref != own:
            return ref
    return f"{own}-{r['part_no']}"

def node_parent_ref(node):
    p = node.parent
    if p is None or p.row is None:
        return (ROOT_DRAW, ROOT_NAME)
    return (node_code(p), p.row['name'])

# rule 1+5+8/9/10: 材质
PROFILE_KEYS = ('角钢', '方钢', '槽钢', '工字钢', 'H型钢', '扁钢')

def classify_material(r):
    """return 'plate' | 'pipe' | 'bar' | 'profile' | None"""
    name = r['name']
    if '钢管' in name or '焊管' in name or '无缝管' in name:
        return 'pipe'
    if '圆钢' in name:
        return 'bar'
    if '钢板' in name:
        return 'plate'
    if any(k in name for k in PROFILE_KEYS):
        return 'profile'
    spec = r['spec']
    if re.match(r'^Ø\d+(?:\.\d+)?x\d+(?:\.\d+)?', spec):
        return 'pipe'
    return None

def plate_thickness(spec):
    nums = [float(m) for m in re.findall(r'\d+(?:\.\d+)?', spec)]
    if not nums:
        return None
    t = min(nums)
    return f"t{t:g}"

def spec_no_length(spec):
    """材质列里不保留长度 L=xxx"""
    return re.sub(r'\s*L=\d+(?:\.\d+)?', '', spec).strip()

def pipe_spec(r):
    spec = r['spec']
    if spec and not spec.startswith('Ø') and re.match(r'\d+x', spec):
        spec = 'Ø' + spec
    return spec

def material_of(node):
    r = node.row
    if r['ares']:                       # 买方供货：不填写
        return ''
    # 非标准件：板材/管材/型材优先（名称判断优先于分类号，防误判标准件）
    cls = classify_material(r)
    if r['material']:
        if cls == 'plate':
            t = plate_thickness(r['spec'])
            return f"钢板{t}/{r['material']}" if t else r['material']
        if cls == 'pipe':
            return f"圆管{spec_no_length(pipe_spec(r))}/{r['material']}"
        if cls == 'bar':
            return f"圆棒{spec_no_length(r['spec'])}/{r['material']}"
        if cls == 'profile':
            base = r.get('name_base', r['name'])
            return f"{base}{spec_no_length(r['spec'])}/{r['material']}"
    # 标准件
    if is_std_part(r):
        parts = []
        if r['spec']:
            parts.append(r['spec'])
        if r['std']:
            s = re.sub(r'图[:：]?|（|）|\(|\)', ' ', r['std'])
            s = re.sub(r'\s+', ' ', s).strip()
            if s:
                parts.append(s)
        if r['material']:
            parts.append(normalize_mat(r['material']))
        return '/'.join(parts)
    return r['material']

# rule 4+5: 备注 = 规格/设备PI号列 (Ares 行留空)
def remark_of(node):
    r = node.row
    if r['ares']:
        return ''
    return r['spec']

def weight_per(node):
    r = node.row
    if r['weight'] is None:
        return None
    q = r['qty'] or 1
    return round(r['weight'] / q, 4)

def total_w(node):
    r = node.row
    if r['weight'] is None:
        return None
    return round(r['weight'] * node.mult, 2)

# ---------- build excel ----------
wb = openpyxl.load_workbook(TEMPLATE)
ws = wb['模板表单']
ws['C1'] = f"{ROOT_CONTRACT} {ROOT_NAME}（{ROOT_DRAW}）"

thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT = Font(name='宋体', size=10)
FONT_B = Font(name='宋体', size=10, bold=True)
ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

r0 = 3
row_of = {id(node): r0 + i for i, node in enumerate(flat)}
for i, node in enumerate(flat):
    r = node.row
    xr = r0 + i
    parent_ref, parent_name = node_parent_ref(node)
    # K(单台总重)、Q/R 为标准件公式列：Q/R 保留模板公式；K 父件=子件之和，叶子=G*J
    data = {
        1: node.number,              # A 序号
        2: node_code(node),          # B 代号
        3: r['name'],                # C 名称
        4: parent_ref,               # D 所属代号
        5: parent_name,              # E 所属部件
        6: r['qty'],                 # F 所属数量
        7: r['qty'] * node.mult,     # G 单台数量
        9: material_of(node),        # I 材质
        10: weight_per(node),        # J 单件重量
        14: remark_of(node),         # N 备注
    }
    bold = bool(node.children)
    for ci in range(1, 20):
        c = ws.cell(xr, ci)
        c.font = FONT_B if bold else FONT
        c.border = border
        c.alignment = ALIGN
        if ci in data:
            c.value = data[ci]
            if ci in (6, 7):
                c.number_format = '0'
            elif ci == 10:
                c.number_format = '0.0000'
    # K 列（单台总重）：父件 = 直接子件 K 之和；叶子 = 单台数量×单件重量
    if node.children:
        ws.cell(xr, 11).value = '=SUM(' + ','.join(f'K{row_of[id(c)]}' for c in node.children) + ')'
    else:
        ws.cell(xr, 11).value = f'=G{xr}*J{xr}'

# rule 3: no colors - clear all fills on data rows (leave template formulas intact)
for xr in range(3, ws.max_row + 1):
    for ci in range(1, 20):
        c = ws.cell(xr, ci)
        if c.fill and c.fill.patternType:
            c.fill = PatternFill()

os.makedirs(OUT_DIR, exist_ok=True)
out_path = os.path.join(OUT_DIR, OUT_FILE)
try:
    wb.save(out_path)
    print('Saved:', out_path)
except PermissionError:
    alt = os.path.join(OUT_DIR, OUT_FILE[:-5] + '（新）.xlsx')
    wb.save(alt)
    print(f'注意: 目标文件正被占用(可能WPS打开)，已另存为:\n  {alt}')
print('Total rows:', len(flat))

# ---------- print review ----------
def fmt(v):
    return '' if v is None else v

print(f"{'序号':<8}{'代号':<18}{'名称':<26}{'所属代号':<16}{'所属部件':<20}{'F':>5}{'G':>6}  {'材质':<38}{'J':>8}{'K':>10} 备注")
for node in flat:
    r = node.row
    pref, pname = node_parent_ref(node)
    print(f"{node.number:<8}{node_code(node):<18}{r['name']:<26}{pref:<16}{pname:<20}{r['qty']:>5}{r['qty']*node.mult:>6}  {material_of(node):<38}{fmt(weight_per(node)):>8}{fmt(total_w(node)):>10} {remark_of(node)}")
