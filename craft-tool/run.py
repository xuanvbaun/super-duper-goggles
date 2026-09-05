# -*- coding: utf-8 -*-
"""工艺清单生成程序（可移植版）
用法:
    python run.py "图纸材料表PDF所在文件夹"
    或直接双击 运行.bat 后输入文件夹路径

流程: 扫描源文件夹(一个或多个顶层PDF + 分/子件PDF) -> OCR -> 表格重建 -> 生成工艺清单Excel
输出: 单根清单沿用设备名称；多个独立顶层清单合并为同文件夹的一份“总清单”。
"""
import os, sys, re, json, hashlib, shutil
from collections import defaultdict, Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(SCRIPT_DIR, '工艺清单标准模版-2026.6.30(1).xlsx')
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')

INT_RE = re.compile(r'^(\d+)')
FLOAT_RE = re.compile(r'\d+\.\d+|\d+')
DRAW_RE = re.compile(r'[A-Z0-9]{2,6}-[A-Z0-9]{2,6}-\d{2,4}(?:-\d{2,4})?')
JUNK_RE = re.compile(r'模板文件|第\d+页|^\d{4}/\d+/\d+|26X-R\d+|26X-k\d+|26X-K\d+')

# 表格列边界（PDF 页面统一渲染为 1685x1192）
# 旧模板：数量列在 x≈245；新模板（Q300 销钉焊接图、Q281 变体）：左列左移/右列右移
BOUNDS_A = [
    ('seq', 245), ('qty', 305), ('unit', 340), ('name', 570), ('spec', 950),
    ('std', 1160), ('material', 1315), ('weight', 1415), ('supplier', 1470), ('cat', 99999),
]
BOUNDS_B = [
    # 阿瑞斯新版表格的“名称/规格”竖线约在 x=500；旧值565会把规格栏开头
    # （如“带法兰密封垫”“Ø354/Ø284x4”）错误拼进名称。
    ('seq', 210), ('qty', 280), ('unit', 315), ('name', 500), ('spec', 950),
    ('std', 1160), ('material', 1330), ('weight', 1440), ('supplier', 1510), ('cat', 99999),
]

def pick_bounds(lines, h_yc):
    """按表头'数量'列的位置选择模板列边界（旧 x≈245 / 新 x≈215）"""
    for l in lines:
        if abs(l[1] - h_yc) <= 40 and l[6] == '数量':
            return BOUNDS_B if l[0] < 240 else BOUNDS_A
    return BOUNDS_A

def col_of(xc, bounds=BOUNDS_A):
    for cn, b in bounds:
        if xc < b:
            return cn
    return 'cat'

PROFILE_KEYS = ('角钢', '方钢', '空心钢', '槽钢', '工字钢', 'H型钢', '扁钢')

# ============================== 基础工具 ==============================
def parse_int(t):
    m = INT_RE.match(t or '')
    return int(m.group(1)) if m else None

def parse_float(t):
    m = FLOAT_RE.search(t or '')
    return float(m.group(0)) if m else None

def clean_spec(t):
    t = re.sub(r"(?<=\d)'(?=\d)", '1', t)  # OCR 常把数字间的 1 读成撇号（如 L=311'5 → L=3115）
    t = re.sub(r'口(?=\d)|(?<=\d)口', '□', t).replace('?', '')  # 方钢符号□(OCR常读成口)，中文"口"不替换
    t = re.sub(r'(?<=\d):(?=\d)', '.', t)       # 冒号误读小数点（2:8→2.8）
    t = re.sub(r'(?<=\D)Q(?=\d)', 'Ø', t)       # Q 误读 Ø（Q26.9→Ø26.9）
    t = re.sub(r',(?=\s*\d)', ' ', t)           # 逗号后接数字是噪声（DN20, 026.9→DN20 026.9）
    t = re.sub(r'(^|[\s,])(==+|b=|=)(?=\d)', r'\1L=', t)  # =/==/b= 误读 L=（=1205→L=1205）
    t = re.sub(r'<=(?=\d)', 'L=', t)                  # <= 误读 L=（<=200→L=200）
    t = re.sub(r'K=(\d+)(?![°度]|\.\d|\d)', r'L=\1', t)   # K 误读 L（K=40→L=40，保留销轴 K=120°）
    t = re.sub(r'L=(\d+)Q(?=\D|$)', r'L=\g<1>0', t)      # L=173Q→L=1730（Q误读0）
    t = re.sub(r'L=(\d)\.(\d)(\d)', r'L=\1\2\3', t)      # L=8.50→L=850, L=1.070→L=1070
    t = t.replace('ⅡI', 'Ⅱ')
    t = t.replace('平热', '平垫').replace('弹热', '弹垫')
    t = re.sub(r'(?<=\d)\s+(?=\d(?:\D|$))', '.', t)  # 3 5→3.5（小数点漏识别）
    t = re.sub(r'^0(?=\d{3,4}/)', 'Ø', t)              # 01100/Ø817→Ø1100/Ø817
    t = t.replace('L=1.6', 'L=16')
    t = re.sub(r'(?<=\d)X(?=\d)', 'x', t)
    ls = re.findall(r'L=\d+(?:\.\d+)?', t)
    rest = re.sub(r'L=\d+(?:\.\d+)?', '', t)
    rest = re.sub(r'(?<![\d.])0(\d{2,3})(?![\d])', r'Ø\1', rest)
    out = re.sub(r'\s+', ' ', rest).strip()
    return (out + ' ' + ' '.join(ls)).strip()

def clean_name(t):
    t = t.replace('0型', 'O型')
    t = t.replace('层步梯组装图', '二层步梯组装图')   # OCR 漏"二"字
    t = re.sub(r'\(带\s*侧喷口[）)]', '(带侧喷口)', t)
    t = re.sub(r'（带\s*侧喷口[）)]', '(带侧喷口)', t)
    t = t.replace('烟肉', '烟囱').replace('烟卤', '烟囱')
    return t.strip()

def clean_std(t):
    t = t.replace('7Q.1', '70.1')
    t = re.sub(r'^[>＞]+', '', t)
    return t.strip()

def normalize_mat(t):
    t = t.replace('（', '(').replace('）', ')')
    return re.sub(r'\s+', '', t)

def normalize_drawing(d):
    return re.sub(r'-[A-Z]$', '', d)

def refs_of(r):
    return re.findall(DRAW_RE, r.get('std', ''))

def line_center(box):
    xs = [p[0] for p in box]; ys = [p[1] for p in box]
    return (min(xs)+max(xs))/2, (min(ys)+max(ys))/2

def line_box(box):
    xs = [p[0] for p in box]; ys = [p[1] for p in box]
    return min(xs), min(ys), max(xs), max(ys)

# ============================== 源文件扫描 ==============================
def drawing_from_filename(f):
    m = re.search(DRAW_RE, f)
    if m:
        return m.group(0)
    return normalize_drawing(re.sub(r'[（(].*?[)）]', '', f).strip())

def scan_sub_pdfs(folder):
    """扫描一个文件夹里的子件PDF，返回 {规范化图号: 路径}"""
    out = {}
    if not os.path.isdir(folder):
        return out
    for f in sorted(os.listdir(folder)):
        if not f.lower().endswith('.pdf'):
            continue
        d = drawing_from_filename(f)
        if d in out and re.search(r'Q281', f):
            continue  # 已有同图号PDF时忽略 Q281 变体（主清单引用 Q280 基础版）
        out[d] = os.path.join(folder, f)
    return out

def find_project(source_dir):
    """返回 (兼容主PDF, 可展开PDF映射, 顶层PDF列表)。

    顶层PDF不能只挑第一份：它们可能是多个独立根清单，也可能互相引用形成
    多棵装配树。第一份仅保留为旧流程的缓存入口，最终根清单由引用关系决定。
    """
    source_dir = os.path.abspath(source_dir)
    if not os.path.isdir(source_dir):
        print(f'错误: 文件夹不存在 -> {source_dir}')
        sys.exit(1)

    pdfs_top = sorted(f for f in os.listdir(source_dir) if f.lower().endswith('.pdf'))
    sub_pdfs = scan_sub_pdfs(os.path.join(source_dir, '分'))
    # 跨项目回退：在源文件夹的上级目录里查找同图号的子件PDF（如 A300 引用 GDL130-Q300-20）
    # 本项目 分 优先，兄弟项目只做缺失回退（setdefault 不覆盖已存在的同图号）
    parent = os.path.dirname(source_dir)
    if os.path.isdir(parent):
        for d in os.listdir(parent):
            dp = os.path.join(parent, d)
            if not os.path.isdir(dp) or d == os.path.basename(source_dir):
                continue
            if d == '分':  # 同级的分文件夹本身就是另一个项目的子件目录
                ext = scan_sub_pdfs(dp)
            else:          # 兄弟项目文件夹内的 分
                ext = scan_sub_pdfs(os.path.join(dp, '分'))
            for k, v in ext.items():
                sub_pdfs.setdefault(k, v)

    main_pdf = None
    if pdfs_top:
        cand = [f for f in pdfs_top if '总图' in f]
        main_pdf = os.path.join(source_dir, cand[0]) if cand else os.path.join(source_dir, pdfs_top[0])
    if main_pdf is None:
        print('错误: 源文件夹中没有找到 PDF 文件')
        sys.exit(1)
    top_paths = [os.path.join(source_dir, f) for f in pdfs_top]
    # 其余顶层PDF也必须进入OCR和子图映射。之后再通过引用图构建真正的根清单，
    # 可同时支持“多棵独立装配树”和“若干完全无层级PDF”。
    main_key = os.path.normcase(os.path.abspath(main_pdf))
    for p in top_paths:
        if os.path.normcase(os.path.abspath(p)) == main_key:
            continue
        sub_pdfs.setdefault(normalize_drawing(drawing_from_filename(os.path.basename(p))), p)
    return main_pdf, sub_pdfs, top_paths

# ============================== OCR 步骤 ==============================
def do_ocr(pdf_path, cache_dir, name):
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR
    print(f'  [OCR] {os.path.basename(pdf_path)} ...')
    try:
        pdf = pdfium.PdfDocument(pdf_path)
    except Exception:
        print(f'  [跳过] 无法读取PDF（可能被WPS损坏）: {os.path.basename(pdf_path)}')
        return None
    ocr = RapidOCR()
    out = {'file': pdf_path, 'pages': []}
    for i in range(len(pdf)):
        img = pdf[i].render(scale=2.0).to_pil()
        tmp = os.path.join(cache_dir, f'{name}_p{i}.png')
        img.save(tmp)
        res, _ = ocr(tmp)
        page = [{'box': it[0], 'text': it[1], 'conf': float(it[2])} for it in (res or [])]
        out['pages'].append(page)
        print(f'  [OCR]   页{i+1}: {len(page)} 行')
        os.remove(tmp)  # 只保留识别结果，不保留临时图片
    with open(os.path.join(cache_dir, name + '.json'), 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False)
    return out

# ============================== 表格重建 ==============================
def meta_of(doc):
    joined = '\n'.join(it['text'] for it in doc['pages'][0])
    m = {}
    for key, pat in [('drawing', r'材料表号[:：]?\s*([A-Za-z0-9\-]+)'),
                     ('qty', r'制造数量：?\s*([0-9]+)'),
                     ('name', r'设备名称：?\s*([^\n]+)'),
                     ('contract', r'合同号：?\s*([^\n]+)')]:
        mm = re.search(pat, joined)
        if mm:
            value = mm.group(1).strip()
            m[key] = clean_name(value) if key == 'name' else value
    return m

def parse_page(page):
    lines = []
    for it in page:
        box, text, conf = it['box'], it['text'], it['conf']
        xc, yc = line_center(box)
        x0, y0, x1, y1 = line_box(box)
        lines.append((xc, yc, y0, y1, x0, x1, text, conf))
    header = next((l for l in lines if l[6] == '名称'), None)
    if header is None:
        header = next((l for l in lines if '名称' in l[6]), None)
    if header is None:
        return None
    h_yc = header[1]
    bounds = pick_bounds(lines, h_yc)
    h_bottom = h_yc + 30
    data = [l for l in lines if l[1] > h_bottom and l[4] > 60 and not JUNK_RE.search(l[6])]
    if not data:
        return []
    data.sort(key=lambda l: l[1])
    bands, cur = [], [data[0]]
    for l in data[1:]:
        if l[1] - cur[-1][1] < 13.5:
            cur.append(l)
        else:
            bands.append(cur); cur = [l]
    bands.append(cur)
    rows = []
    for band in bands:
        rec = {'y': (min(l[1] for l in band) + max(l[1] for l in band)) / 2, 'cols': defaultdict(list)}
        for l in band:
            rec['cols'][col_of(l[0], bounds)].append((l[6], l[7], l[2]))
        rows.append(rec)
    return rows

def build_tables(cache_dir, only_names=None):
    """读取缓存目录里的 OCR 结果重建表格。
    only_names: 只读取这些缓存名（本次实际扫描到的 PDF），防止已删除子件的旧缓存污染结果。"""
    results = {}
    for f in sorted(os.listdir(cache_dir)):
        if not f.endswith('.json'):
            continue
        name = f[:-5]
        if only_names is not None and name not in only_names:
            continue
        with open(os.path.join(cache_dir, f), encoding='utf-8') as fh:
            doc = json.load(fh)
        records = []
        total_w = None
        for pi, page in enumerate(doc['pages']):
            rows = parse_page(page)
            if rows is None:
                continue
            for r in rows:
                rec = {}
                for cn in ('seq','qty','unit','name','spec','std','material','weight','supplier','cat'):
                    if cn in r['cols']:
                        vals = sorted(r['cols'][cn], key=lambda v: v[2])
                        text = re.sub(r'\s+', ' ', ' '.join(v[0] for v in vals)).strip()
                        if text:
                            rec[cn] = (text, min(v[1] for v in vals))
                for k in rec:
                    m = re.search(r'总重[：:]\s*([\d.]+)', rec[k][0])
                    if m:
                        total_w = float(m.group(1))   # 页脚总重（最后出现的为准）
                if len(rec) == 1 and 'weight' in rec:
                    continue
                if any('总重' in rec[k][0] for k in rec):
                    continue
                if set(rec.keys()) <= {'supplier', 'cat'}:
                    continue
                rec['_page'] = pi
                rec['_y'] = r['y']
                records.append(rec)
        results[name] = {'file': doc['file'], 'records': records, 'meta': meta_of(doc),
                         'total_weight': total_w, 'pages': doc.get('pages', [])}
    return results

# ============================== PART/WEIGHT REREAD ==============================
WEIGHT_RE = re.compile(r'^\d+(?:\.\d{1,4})?$')
PART_RE = re.compile(r'^\d+(?:\.\d+)*$')
QTY_RE = re.compile(r'\d+(?:\.\d+)?')

def _clean_weight(val):
    s = (val or '').strip().replace(',', '.')
    s = re.sub(r'\s+', '', s)
    if WEIGHT_RE.fullmatch(s):
        try:
            return float(s)
        except ValueError:
            return None
    return None

def _clean_part_no(val):
    s = re.sub(r'\s+', '', (val or '').strip())
    return s if PART_RE.fullmatch(s) else None

def _clean_qty(val):
    """数量列只接受唯一的正数；允许小数数量及 OCR 带出的相邻单位文字。"""
    nums = QTY_RE.findall((val or '').strip())
    if len(nums) != 1:
        return None
    q = float(nums[0])
    if not (0 < q <= 100000):
        return None
    return int(q) if q.is_integer() else q

def _page_lines(page):
    lines = []
    for it in page or []:
        box, text, conf = it['box'], it['text'], it['conf']
        xc, yc = line_center(box)
        x0, y0, x1, y1 = line_box(box)
        lines.append((xc, yc, y0, y1, x0, x1, text, conf))
    return lines

def _find_header(lines, keywords):
    exact = [l for l in lines if any(l[6].strip() == k for k in keywords)]
    if exact:
        return min(exact, key=lambda l: l[1])
    fuzzy = [l for l in lines if any(k in l[6].replace(' ', '') for k in keywords)]
    if fuzzy:
        return min(fuzzy, key=lambda l: l[1])
    return None

def _column_window(page, kind):
    lines = _page_lines(page)
    if not lines:
        return None

    name_h = _find_header(lines, ('名称',))
    h_yc = name_h[1] if name_h else min(l[1] for l in lines)
    bounds = pick_bounds(lines, h_yc)

    if kind == 'weight':
        cur = _find_header(lines, ('重量',))
        left_h = _find_header(lines, ('材料/物料编码', '物料编码', '材料'))
        right_h = _find_header(lines, ('供货',))
        if cur:
            cx = cur[0]
            left = (left_h[0] + cx) / 2 if left_h and left_h[0] < cx else cx - 70
            right = (cx + right_h[0]) / 2 if right_h and right_h[0] > cx else cx + 70
            return max(0, left - 8), right + 8
        prev = 0
        for cn, b in bounds:
            if cn == 'weight':
                return max(0, prev - 10), b + 10
            prev = b
        return None

    if kind == 'seq':
        # 表头常被 OCR 拆成“件”/“号”两行，不能只查完整“件号”。
        cur = _find_header(lines, ('件号', '件'))
        left_h = _find_header(lines, ('修改', '修'))
        right_h = _find_header(lines, ('数量',))
        if not cur:
            return None
        cx = cur[0]
        left = (left_h[0] + cx) / 2 if left_h and left_h[0] < cx else cx - 35
        right = (cx + right_h[0]) / 2 if right_h and right_h[0] > cx else cx + 35
        return max(0, left + 2), right - 2

    if kind == 'qty':
        cur = _find_header(lines, ('数量',))
        left_h = _find_header(lines, ('件号',))
        right_h = _find_header(lines, ('单位',))
        if not cur:
            return None
        cx = cur[0]
        left = (left_h[0] + cx) / 2 if left_h and left_h[0] < cx else cx - 28
        right = (cx + right_h[0]) / 2 if right_h and right_h[0] > cx else cx + 28
        return max(0, left + 2), right - 2

    return None

def _ocr_row_cell(img6, x0, x1, row_y, ocr, suffix):
    import tempfile
    from PIL import Image
    k = 3.0
    y0 = max(0, row_y - 22)
    y1 = row_y + 22
    crop = img6.crop((int(x0 * k), int(y0 * k), int(x1 * k), int(y1 * k)))
    if crop.width <= 0 or crop.height <= 0:
        return []
    crop = crop.resize((crop.width * 3, crop.height * 3), Image.LANCZOS)
    tmp = tempfile.NamedTemporaryFile(prefix='craft_', suffix=suffix, delete=False)
    tmp.close()
    try:
        crop.save(tmp.name)
        res, _ = ocr(tmp.name)
        return [it[1].strip() for it in (res or []) if it[1].strip()]
    finally:
        try:
            os.remove(tmp.name)
        except OSError:
            pass

def reread_missing_part_numbers(tables):
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR

    ocr = RapidOCR()
    filled = changed = 0
    for t in tables.values():
        pages = t.get('pages') or []
        if not pages:
            continue
        missing_by_page = {}
        for rec in t['records']:
            raw = rec.get('seq', ('', 0))[0].strip() if 'seq' in rec else ''
            missing_by_page.setdefault(rec.get('_page', 0), []).append(rec)
        if not missing_by_page:
            continue
        try:
            pdf = pdfium.PdfDocument(t['file'])
        except Exception:
            continue
        for pi, recs in missing_by_page.items():
            if pi >= len(pages) or pi >= len(pdf):
                continue
            win = _column_window(pages[pi], 'seq')
            if not win:
                continue
            img6 = pdf[pi].render(scale=6.0).to_pil()
            for rec in recs:
                texts = _ocr_row_cell(img6, win[0], win[1], rec['_y'], ocr, '_seq.png')
                vals = []
                for text in texts:
                    direct = _clean_part_no(text)
                    if direct:
                        vals.append(direct)
                        continue
                    found = re.findall(r'\d+(?:\.\d+)*', text)
                    if len(found) == 1 and PART_RE.fullmatch(found[0]):
                        vals.append(found[0])
                vals = [x for x in vals if x]
                uniq = list(dict.fromkeys(vals))
                if len(uniq) == 1:
                    old = _clean_part_no(raw)
                    rec['seq'] = (uniq[0], 1.0)
                    rec['_seq_reread'] = True
                    if old is None:
                        filled += 1
                    elif old != uniq[0]:
                        rec['_seq_original'] = old
                        changed += 1
    print(f"Part-number reread: corrected {changed}, filled {filled}")

def reread_missing_weights(tables):
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR

    ocr = RapidOCR()
    filled = 0
    for t in tables.values():
        pages = t.get('pages') or []
        if not pages:
            continue
        missing_by_page = {}
        for rec in t['records']:
            if 'weight' not in rec:
                missing_by_page.setdefault(rec.get('_page', 0), []).append(rec)
        if not missing_by_page:
            continue
        try:
            pdf = pdfium.PdfDocument(t['file'])
        except Exception:
            continue
        for pi, recs in missing_by_page.items():
            if pi >= len(pages) or pi >= len(pdf):
                continue
            win = _column_window(pages[pi], 'weight')
            if not win:
                continue
            img6 = pdf[pi].render(scale=6.0).to_pil()
            for rec in recs:
                texts = _ocr_row_cell(img6, win[0], win[1], rec['_y'], ocr, '_weight.png')
                vals = [_clean_weight(x) for x in texts]
                vals = [x for x in vals if x is not None]
                uniq = list(dict.fromkeys(vals))
                if len(uniq) == 1:
                    rec['weight'] = (str(uniq[0]), 1.0)
                    rec['_weight_reread'] = True
                    filled += 1
    if filled:
        print("Weight reread filled:", filled)

def reread_quantities(tables):
    """只对低倍率完全没有读出数字的数量格做高倍率补读。
    已有数量即使与高倍率结果冲突也不自动覆盖，冲突交给复核报告；
    避免把清晰的“1”在放大后误读成“7”。"""
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR

    ocr = RapidOCR()
    filled = 0
    for t in tables.values():
        pages = t.get('pages') or []
        if not pages:
            continue
        by_page = defaultdict(list)
        for rec in t['records']:
            raw = rec.get('qty', ('', 0))[0] if 'qty' in rec else ''
            if _clean_qty(raw) is None:
                by_page[rec.get('_page', 0)].append(rec)
        if not by_page:
            continue
        try:
            pdf = pdfium.PdfDocument(t['file'])
        except Exception:
            continue
        for pi, recs in by_page.items():
            if pi >= len(pages) or pi >= len(pdf):
                continue
            win = _column_window(pages[pi], 'qty')
            if not win:
                continue
            img6 = pdf[pi].render(scale=6.0).to_pil()
            for rec in recs:
                texts = _ocr_row_cell(img6, win[0], win[1], rec['_y'], ocr, '_qty.png')
                vals = [_clean_qty(x) for x in texts]
                vals = [x for x in vals if x is not None]
                uniq = list(dict.fromkeys(vals))
                if len(uniq) != 1:
                    continue
                new = uniq[0]
                rec['qty'] = (str(new), 1.0)
                rec['_qty_reread'] = True
                filled += 1
    print(f"Quantity reread: filled {filled}; existing values were not overwritten")

# ============================== 已确认修正（外置文件） ==============================
def load_corrections():
    p = os.path.join(SCRIPT_DIR, 'rules', 'corrections.json')
    try:
        with open(p, encoding='utf-8') as f:
            return json.load(f).get('corrections', [])
    except Exception as e:
        print(f'[警告] 读取修正文件失败（将跳过已确认修正）: {p} ({e})')
        return []

CORRECTIONS = load_corrections()

def apply_corrections(records, corrections):
    """按 corrections.json 的规则修正记录。
    匹配字段：drawing(图号) / drawings(图号列表) / name / spec_contains / spec_eq / qty_eq / weight_gt / weight_is_null；
    匹配后把 set 里的字段改为新值。"""
    for r in records:
        for c in corrections:
            if c.get('drawing') and c['drawing'] != r['doc_drawing']:
                continue
            if c.get('drawings') and r['doc_drawing'] not in c['drawings']:
                continue
            if c.get('name') and c['name'] != r['name']:
                continue
            if c.get('spec_contains') and not all(s in r['spec'] for s in c['spec_contains']):
                continue
            if c.get('spec_eq') is not None and r['spec'] != c['spec_eq']:
                continue
            if c.get('qty_eq') is not None and r['qty'] != c['qty_eq']:
                continue
            if c.get('weight_gt') is not None and not (r['weight'] is not None and r['weight'] > c['weight_gt']):
                continue
            if c.get('weight_is_null') and r['weight'] is not None:
                continue
            for k, v in c['set'].items():
                r[k] = v

def capture_total_weights(tables):
    """补全各文档页脚"总重：XXX kg"：低倍率没捕获到的，对最后一页底部区域 scale6 复读。
    返回 {doc名: 总重}（仅补缺失的，已有值保留）。"""
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR
    from PIL import Image
    T_RE = re.compile(r'总重[：:]\s*([\d.]+)')
    ocr = RapidOCR()
    out = {}
    for name, t in tables.items():
        if t.get('total_weight') is not None:
            continue
        try:
            pdf = pdfium.PdfDocument(t['file'])
        except Exception:
            continue
        try:
            pi = len(pdf) - 1
            img = pdf[pi].render(scale=6.0).to_pil()
            w, h = img.size
            crop = img.crop((int(0.55 * w), int(0.60 * h), int(0.92 * w), int(0.99 * h)))
            crop = crop.resize((crop.width * 2, crop.height * 2), Image.LANCZOS)
            crop.save('_tw.png')
            res, _ = ocr('_tw.png')
            os.remove('_tw.png')
            txt = ''.join(it[1] for it in (res or []))
            m = T_RE.search(txt)
            if m:
                out[name] = float(m.group(1))
        except Exception:
            continue
    return out

# ============================== 行处理 ==============================
def process_doc(t):
    doc_drawing = normalize_drawing(t['meta'].get('drawing', ''))
    out = []
    for rec in t['records']:
        r = {
            'seq':   parse_int(rec['seq'][0]) if 'seq' in rec else None,
            'seq_raw': rec['seq'][0].strip() if 'seq' in rec else '',   # PDF原件号原文（保留 2.1 分层件号）
            'qty':   _clean_qty(rec['qty'][0]) if 'qty' in rec else 1,
            'unit':  rec['unit'][0].replace('祥', '件') if 'unit' in rec else '件',
            'name':  clean_name(rec['name'][0]) if 'name' in rec else '',
            'spec':  clean_spec(rec['spec'][0]) if 'spec' in rec else '',
            'std':   clean_std(rec['std'][0]) if 'std' in rec else '',
            'material': clean_spec(rec['material'][0]) if 'material' in rec else '',
            'weight':   parse_float(rec['weight'][0]) if 'weight' in rec else None,
            'supplier': rec['supplier'][0].strip() if 'supplier' in rec else '',
            'cat':      rec['cat'][0].strip() if 'cat' in rec else '',
            'doc_drawing': doc_drawing,
            'ares': False,
        }
        # 新版阿瑞斯表格中“重量”和“供货”两格间距很窄，OCR 经常把它们合成
        # 一个文本框（如“23.21 Ares”或“0.12”），其中心点会落入供货列。
        # 当重量栏为空时，从这个合并框中回收重量；供货标记本身继续保留。
        if r['weight'] is None and r['supplier']:
            weight_match = re.search(r'(?<![A-Za-z0-9])\d+\.\d+(?![A-Za-z0-9])', r['supplier'])
            if weight_match:
                r['weight'] = float(weight_match.group(0))
                r['supplier'] = re.sub(r'\s+', ' ',
                                       r['supplier'][:weight_match.start()] + ' ' +
                                       r['supplier'][weight_match.end():]).strip()
        r['ares'] = 'Ares' in r['supplier']
        # OCR 偶尔把规格栏开头的管径粘到名称末尾，例如：
        # “流体输送用不锈钢无缝钢管021x3.0” + “L=300”。
        # 将末尾尺寸移回规格栏，避免输出成“圆管/材质”。
        leaked_pipe_spec = re.search(r'([0QØ]\d+(?:\.\d+)?x\d+(?:\.\d+)?)$', r['name'])
        if leaked_pipe_spec and any(k in r['name'] for k in ('钢管', '焊管', '无缝管')):
            leaked = leaked_pipe_spec.group(1)
            if leaked.startswith(('0', 'Q')):
                leaked = 'Ø' + leaked[1:]
            r['name'] = r['name'][:-len(leaked_pipe_spec.group(1))].rstrip()
            r['spec'] = clean_spec(f"{leaked} {r['spec']}")
        if not r['name'] and not r['spec'] and not r['std']:
            continue
        # 纯说明行不是物料，不能混入 Excel（例如“注：镜像制造，各做一件”）。
        if (r['name'].startswith(('注：', '注:')) and not r['spec'] and not r['std']
                and not r['material'] and r['weight'] is None):
            continue
        if r['qty'] is None:
            r['qty'] = 1
        # 常见 OCR 修正
        r['std'] = r['std'].replace('GDL465-', 'GDL165-')
        if r['name'] == '流体输送用无缝钢管' and '146x4.0' in r['spec']:
            r['spec'] = r['spec'].replace('146x4.0', 'Ø146x4.0')   # 缺Ø(复核确认)
        # 管状卷制件的直径符号位于规格开头时，OCR 常把 Ø 读成 Q；
        # clean_spec 不能全局替换开头 Q，否则会破坏 Q235-A 这类材质牌号。
        if ('卷管' in r['name'] or '钢筒' in r['name']) and re.match(r'^Q\d+[xX]', r['spec']):
            r['spec'] = 'Ø' + r['spec'][1:]
        if r['name'] == '耐热钢板':
            if 'Q154' in r['spec']:
                r['spec'] = '2×Ø154'          # Q为0误读(复核确认)
            elif '2×154' in r['spec']:
                r['spec'] = '2×Ø154'          # 缺Ø(复核确认)
        if r['name'] == '时效炉风机型线2中间风箱' and r['qty'] == 9:
            r['qty'] = 6                 # OCR 把6误读为9(高倍率复核确认, 与子件图纸制造数量一致)
        if r['name'] == '陶瓷纤维方编绳':
            r['material'] = ''           # 材质格在PDF中为空白
        if r['name'] == '耐热钢板' and '2xØ90' in r['spec']:
            r['qty'] = 1                 # 数量格在PDF中为空白
        r['material'] = re.sub(r'\s*26X\S*', '', r['material']).strip()   # 合同号渗出噪声(复核确认)
        r['spec'] = re.sub(r'^\d+\s+(?=90E\(L\))', '', r['spec'])          # 90°弯头前的孤立数字噪声
        out.append(r)
    apply_corrections(out, CORRECTIONS)   # 已人工确认的修正（rules/corrections.json）
    assign_part_no(out)
    return out

SEQ_RE = re.compile(r'^\d+(\.\d+)*$')   # 合法件号格式：45 / 2.1 / 19.2

def assign_part_no(records):
    """件号 = PDF 原件号（字符串，保留 2.1 这类分层件号）。
    不因行数/重复/缺失重新编号（图纸件号是图纸标识，程序无权重编）；
    缺失或格式明显异常（如 OCR 乱码）时置 None，由复核报告标记。"""
    for r in records:
        raw = r.get('seq_raw', '')
        r['part_no'] = raw if SEQ_RE.match(raw) else None

STD_NAME_RE = re.compile(r'螺栓|螺钉|螺母|垫圈|开口销|销轴|铆钉|挡圈|键|轴承|油杯')

def is_std_part(r):
    if refs_of(r):
        return False  # 有子图号引用的是自制件，不是标准件（分类号列常被OCR误读成16）
    if r['cat'] in ('16', '17', '7') and '编绳' not in r['name']:
        return True
    # 分类号列常被 OCR 读坏（如 116 2 4 / 空），用"标准件名称+GB/T/JB/T标准号"兜底
    return (bool(STD_NAME_RE.search(r['name'])) and bool(re.search(r'GB/?T|JB/?T', r['std']))
            and '编绳' not in r['name'])

def classify_material(r):
    name = r['name']
    if any(k in name for k in ('钢管', '焊管', '无缝管', '卷管', '钢筒')):
        return 'pipe'
    if '圆钢' in name:
        return 'bar'
    if '钢板' in name:
        return 'plate'
    if any(k in name for k in PROFILE_KEYS):
        return 'profile'
    if re.match(r'^Ø\d+(?:\.\d+)?x\d+(?:\.\d+)?', r['spec']):
        return 'pipe'
    return None

def plate_thickness(spec):
    nums = [float(m) for m in re.findall(r'\d+(?:\.\d+)?', spec)]
    if not nums:
        return None
    return f"t{min(nums):g}"

def spec_no_length(spec):
    return re.sub(r'\s*L=\d+(?:\.\d+)?', '', spec).strip()

def base_section_spec(spec, cls):
    """材质列只保留型材/管材的基础截面尺寸，不带长度、孔和加工说明。"""
    s = spec_no_length(spec).strip()
    if not s:
        return ''
    if cls == 'profile':
        # 400x102x12.5x18/26-Ø255 -> 400x102x12.5x18
        return re.split(r'[/（(；;，,]', s, maxsplit=1)[0].strip()
    if cls == 'pipe':
        # DN50 Ø60.3x3.8、Ø42x3.0、Ø256x1 等只保留截面尺寸。
        m = re.match(r'(DN\d+(?:\.\d+)?(?:\s+Ø\d+(?:\.\d+)?(?:x\d+(?:\.\d+)?)?)?)', s, re.I)
        if m:
            return m.group(1).strip()
        m = re.match(r'(Ø\d+(?:\.\d+)?(?:x\d+(?:\.\d+)?)?)', s, re.I)
        if m:
            return m.group(1).strip()
        # 钢筒类图纸可能采用 D254-46.5;t=1mm 这种专用截面标注，保留尺寸本身。
        m = re.match(r'(D\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?(?:;\s*t=\d+(?:\.\d+)?mm)?)', s, re.I)
        if m:
            return m.group(1).strip()
        return re.split(r'[/（(；;，,]', s, maxsplit=1)[0].strip()
    if cls == 'bar':
        return re.split(r'[/（(；;，,]', s, maxsplit=1)[0].strip()
    return s

def profile_name(name):
    """型材在材质列只保留通用名称，去掉热轧/冷弯/等边/对称制作等修饰词。"""
    if '空心钢' in name or '方钢' in name:
        return '方管'
    if '槽钢' in name:
        return '槽钢'
    if '角钢' in name:
        return '角钢'
    if 'H型钢' in name:
        return 'H型钢'
    if '工字钢' in name:
        return '工字钢'
    if '扁钢' in name:
        return '扁钢'
    return '型材'

def material_of(r):
    # 用户要求材质列与 PDF“材料/物料编码”栏逐格一致。规格和标准不得再
    # 拼入材质列；PDF 原格为空时保持空白（包括供货件和标准件）。
    return r['material']

# ============================== 高倍率复核 ==============================
VERIFY_COLS_A = {'qty': (240, 310), 'weight': (1310, 1410), 'spec': (560, 960)}
VERIFY_COLS_B = {'qty': (210, 290), 'weight': (1330, 1440), 'spec': (500, 950)}

def norm_cell(cname, t):
    t = (t or '').strip()
    if cname in ('qty', 'weight'):
        m = re.search(r'\d+(?:\.\d+)?', t)
        return m.group(0) if m else ''
    # 规格: 忽略顺序/空格差异, 保留内容差异 —— 比较"数字序列 + 字母符号集合"
    t = t.lower().replace('×', 'x')
    nums = 'N' + '|'.join(sorted(re.findall(r'\d+(?:\.\d+)?', t)))
    chars = 'C' + ''.join(sorted(re.findall(r'[a-zØ]', t)))
    return nums + chars

def verify_cells(pdf_path, doc, cache_dir):
    """对每页的 数量/重量/规格 列做高倍率(scale6)复核。
    返回 (差异列表, 复核单元格总数)"""
    import pypdfium2 as pdfium
    from rapidocr_onnxruntime import RapidOCR
    from PIL import Image
    ocr = RapidOCR()
    try:
        pdf = pdfium.PdfDocument(pdf_path)
    except Exception:
        print(f'  [复核跳过] 无法读取PDF(可能被WPS损坏): {os.path.basename(pdf_path)}')
        return [], 0
    diffs, total = [], 0
    for pi, page in enumerate(doc['pages']):
        lines = []
        for it in page:
            box, text, conf = it['box'], it['text'], it['conf']
            xs = [p[0] for p in box]; ys = [p[1] for p in box]
            lines.append(((min(xs)+max(xs))/2, (min(ys)+max(ys))/2, min(ys), max(ys),
                          min(xs), max(xs), text, conf))
        header = next((l for l in lines if l[6] == '名称'), None)
        if header is None:
            header = next((l for l in lines if '名称' in l[6]), None)
        if header is None:
            continue
        h_yc = header[1]
        bounds = pick_bounds(lines, h_yc)
        verify_cols = VERIFY_COLS_B if bounds is BOUNDS_B else VERIFY_COLS_A
        h_bottom = h_yc + 30
        data = [l for l in lines if l[1] > h_bottom and l[4] > 60 and not JUNK_RE.search(l[6])]
        data.sort(key=lambda l: l[1])
        bands, cur = [], [data[0]] if data else []
        for l in data[1:]:
            if l[1] - cur[-1][1] < 13.5:
                cur.append(l)
            else:
                bands.append(cur); cur = [l]
        if cur:
            bands.append(cur)
        img6 = pdf[pi].render(scale=6.0).to_pil()
        for band in bands:
            y0 = min(l[2] for l in band) - 6
            y1 = max(l[3] for l in band) + 6
            cells = {}
            for l in band:
                cells.setdefault(col_of(l[0], bounds), []).append(l)
            name = ''
            if 'name' in cells:
                name = ' '.join(l[6] for l in sorted(cells['name'], key=lambda l: l[2]))
            for cname, (x0, x1) in verify_cols.items():
                if cname not in cells:
                    continue
                vals = sorted(cells[cname], key=lambda l: l[2])
                orig = ' '.join(l[6] for l in vals)
                total += 1
                crop = img6.crop((int(x0 * 3), int(y0 * 3), int(x1 * 3), int(y1 * 3)))
                crop = crop.resize((crop.width * 3, crop.height * 3), Image.LANCZOS)
                crop.save('_v.png')
                res, _ = ocr('_v.png')
                high = ''.join(it[1] for it in (res or []))
                if norm_cell(cname, orig) != norm_cell(cname, high):
                    diffs.append({'page': pi + 1, 'name': name, 'col': cname,
                                  'orig': orig.strip(), 'high': high.strip()})
    if os.path.exists('_v.png'):
        os.remove('_v.png')
    return diffs, total

def run_verify(doc_name, pdf_path, doc, cache_dir, report):
    """执行复核并写入报告列表, 返回 (差异数, 复核格数)"""
    diffs, total = verify_cells(pdf_path, doc, cache_dir)
    if diffs:
        report.append(f'--- {os.path.basename(pdf_path)} ({len(diffs)} 处差异 / 复核 {total} 格) ---')
        for d in diffs:
            report.append(f'  第{d["page"]}页 | {d["name"][:22]} | {d["col"]}: 原识别={d["orig"]}  高倍率={d["high"]}')
    return diffs, total

# ============================== 层级与生成 ==============================
class Node:
    __slots__ = ('row', 'children', 'level', 'mult', 'parent', 'number')
    def __init__(self, row=None):
        self.row = row
        self.children = []
        self.level = 0
        self.mult = 1
        self.parent = None
        self.number = ''

def build_tree(main_records, sub_map, root_draw=''):
    root = Node()
    # 当前递归链上的图号集合（防循环引用 A→B→C→A 无限递归；同一子图被不同父件
    # 重复引用是合法的，所以不能用全局 visited，只用当前路径）
    active = {root_draw} if root_draw else set()

    def attach(node, records):
        for r in records:
            child = Node(r)
            child.parent = node
            child.level = node.level + 1
            child.mult = node.mult * (node.row['qty'] if node.row else 1)
            node.children.append(child)
            for ref in refs_of(r):
                if ref in sub_map:
                    if ref in active:
                        print(f'  [警告] 检测到循环引用: {ref}（本分支停止展开）')
                        break
                    active.add(ref)
                    attach(child, sub_map[ref])
                    active.discard(ref)
                    break
    attach(root, main_records)
    return root

def entry_aliases(entry):
    return {entry.get('internal', ''), entry.get('filename', '')} - {''}

def select_root_entries(all_entries, top_pdfs):
    """从顶层PDF中选出没有被其他顶层材料表引用的独立根清单。

    返回 (根清单, 顶层清单, 是否使用全量回退)。顶层清单严格沿用文件名排序，
    使多根合并结果稳定。若循环/自引用导致没有普通根，则保留全部顶层清单，
    绝不能只取第一份而遗漏其余材料表。
    """
    entries_by_path = {
        os.path.normcase(os.path.abspath(e['table']['file'])): e
        for e in all_entries
    }
    top_entries = []
    for path in top_pdfs:
        entry = entries_by_path.get(os.path.normcase(os.path.abspath(path)))
        if entry is not None:
            top_entries.append(entry)
    referenced = set()
    for entry in top_entries:
        for rec in entry['records']:
            referenced.update(normalize_drawing(x) for x in refs_of(rec))
    roots = [e for e in top_entries if entry_aliases(e).isdisjoint(referenced)]
    fallback_all = not roots and bool(top_entries)
    return (top_entries if fallback_all else roots), top_entries, fallback_all

def node_code(node):
    r = node.row
    own = r['doc_drawing']
    refs = refs_of(r)
    if refs:
        return normalize_drawing(refs[0])  # 明确引用子图时，代号优先采用引用图号
    if not r['part_no']:
        return own          # 件号缺失/格式异常：代号只保留图号（复核报告已标记）
    return f"{own}-{r['part_no']}"

def weight_per(r):
    if r.get('unit_weight') is not None:
        return r['unit_weight']             # 单元格显示4位，内部保留原始精度
    if r['weight'] is None:
        return None
    return r['weight'] / (r['qty'] or 1)   # 不先四舍五入，防止再乘数量后产生0.0001误差

def detail_weight_total(records):
    """材料表明细重量合计；重量栏是该行数量对应的总重，不再乘数量。"""
    weights = [r['weight'] for r in records if r.get('weight') is not None]
    return round(sum(weights), 4) if weights else None

# ============================== 缓存键 ==============================
def file_key(pdf_path):
    """按 规范化绝对路径 + 大小 + 修改时间 生成缓存键。
    同名 PDF（兄弟项目）不撞名；PDF 更新过（size/mtime 变化）自动换新键重新 OCR。"""
    st = os.stat(pdf_path)
    ident = f"{os.path.normcase(os.path.abspath(pdf_path))}|{st.st_size}|{int(st.st_mtime)}"
    return hashlib.md5(ident.encode('utf-8')).hexdigest()[:12]

def run(source_dir):
    main_pdf, sub_pdfs, top_pdfs = find_project(source_dir)

    # ---- 格式分流：文字型PDF(如瑞士Rowa BOM)走文本提取+翻译路径 ----
    from pdfminer.high_level import extract_text
    try:
        txt = extract_text(main_pdf) or ''
    except Exception:
        txt = ''
    if len(top_pdfs) == 1 and len(txt.strip()) > 200 and not os.path.isdir(os.path.join(source_dir, '分')):
        print('检测到文字型 PDF（Rowa/瑞士格式），走文本提取+翻译路径 ...')
        import swiss_build
        out = swiss_build.build_folder(source_dir)
        if out:
            print(f'\n完成! 输出: {out}')
        return

    print(f'兼容入口PDF: {os.path.basename(main_pdf)}')
    print(f'顶层PDF: {len(top_pdfs)} 个；可展开PDF: {len(sub_pdfs)} 个')

    # 缓存目录（按源文件夹路径哈希，重复运行可跳过OCR）
    key = hashlib.md5(source_dir.encode('utf-8')).hexdigest()[:10]
    cache_dir = os.path.join(DATA_DIR, key)
    os.makedirs(cache_dir, exist_ok=True)

    main_name = 'MAIN_' + file_key(main_pdf)
    main_doc = None
    main_json = os.path.join(cache_dir, main_name + '.json')
    if os.path.exists(main_json):
        with open(main_json, encoding='utf-8') as f:
            main_doc = json.load(f)
        print('(使用已缓存的OCR结果)')
    else:
        main_doc = do_ocr(main_pdf, cache_dir, main_name)
    if main_doc is None:
        print(f'\n错误: 主清单PDF无法读取（可能被WPS损坏）: {os.path.basename(main_pdf)}')
        print('请重新获取这份PDF（不要用WPS打开保存它）后再运行。')
        sys.exit(1)
    docs = {main_name: main_doc}

    # 本项目“分”目录全部保留；兄弟项目只按主表/已加载子表中的实际引用递归加载。
    # 这样仍支持跨项目引用，但不会把同级所有项目混进 OCR 和复核报告。
    source_abs = os.path.normcase(os.path.abspath(source_dir))
    def is_local_pdf(path):
        try:
            return os.path.commonpath([source_abs, os.path.normcase(os.path.abspath(path))]) == source_abs
        except ValueError:
            return False
    wanted = {normalize_drawing(x) for x in DRAW_RE.findall(
        '\n'.join(it['text'] for page in main_doc.get('pages', []) for it in page))}
    pending = [(d, p) for d, p in sub_pdfs.items() if is_local_pdf(p) or normalize_drawing(d) in wanted]
    loaded_paths = set()
    pos = 0
    while pos < len(pending):
        d, p = pending[pos]
        pos += 1
        path_key = os.path.normcase(os.path.abspath(p))
        if path_key in loaded_paths:
            continue
        loaded_paths.add(path_key)
        sub_name = 'SUB_' + file_key(p)
        sj = os.path.join(cache_dir, sub_name + '.json')
        if os.path.exists(sj):
            with open(sj, encoding='utf-8') as f:
                docs[sub_name] = json.load(f)
            print(f'(使用已缓存的OCR结果: {os.path.basename(p)})')
        else:
            sub_doc = do_ocr(p, cache_dir, sub_name)
            if sub_doc is None:
                continue  # 子件PDF损坏则跳过（对应图号不展开）
            docs[sub_name] = sub_doc
        doc_now = docs.get(sub_name, {})
        new_refs = {normalize_drawing(x) for x in DRAW_RE.findall(
            '\n'.join(it['text'] for page in doc_now.get('pages', []) for it in page))}
        wanted.update(new_refs)
        queued = {os.path.normcase(os.path.abspath(x[1])) for x in pending}
        for ref in new_refs:
            ext = sub_pdfs.get(ref)
            if ext and os.path.normcase(os.path.abspath(ext)) not in queued | loaded_paths:
                pending.append((ref, ext))

    # 只读取本次实际扫描到的缓存（防止已删除子件的旧缓存污染结果）
    tables = build_tables(cache_dir, only_names=set(docs.keys()))
    main_t = tables[main_name]

    # ---- 高倍率复读：可靠结果直接回写，复核报告保留变更轨迹 ----
    reread_missing_part_numbers(tables)
    reread_missing_weights(tables)
    reread_quantities(tables)

    # ---- 子清单页脚总重捕获（低倍率缺失时 scale6 复读）----
    extra_totals = capture_total_weights(tables)

    # ---- 高倍率复核：数量/重量/规格 列（图片型OCR路径自动执行）----
    print('\n正在做高倍率复核（数量/重量/规格列）...')
    report = ['高倍率复核报告（数量/重量/规格列）', '=' * 60]
    n_diff = n_total = 0
    for sub_name, t in tables.items():
        doc = docs.get(sub_name)
        if doc is None or (sub_name != main_name and not is_local_pdf(t['file'])):
            continue
        d, tot = run_verify(sub_name, doc['file'], doc, cache_dir, report)
        n_diff += len(d)
        n_total += tot
    print(f'复核完成: 共 {n_total} 格, 发现 {n_diff} 处与原识别不一致')
    rp = os.path.join(source_dir, '复核报告.txt')

    # 所有实际加载的PDF（包括兼容入口PDF）同时按“PDF内部材料表号”和
    # “文件名图号”建立别名。这样第一份PDF即使恰好是子清单，也仍能被
    # 其他顶层根清单正确引用和展开。
    sub_map = {}
    all_entries = []
    drawing_alias_issues = []
    for sub_name, t in tables.items():
        internal = normalize_drawing(t['meta'].get('drawing', ''))
        filename_draw = normalize_drawing(drawing_from_filename(os.path.basename(t['file'])))
        recs = process_doc(t)
        # 文件名/父表引用与PDF内部材料表号冲突时，输出采用文件名/引用图号；
        # 内部号仍记录在复核报告中，不能静默忽略。
        if internal and filename_draw and internal != filename_draw:
            for rec in recs:
                rec['doc_drawing'] = filename_draw
        entry = {'name': sub_name, 'table': t, 'records': recs,
                 'internal': internal, 'filename': filename_draw}
        all_entries.append(entry)
        if internal:
            sub_map[internal] = recs
        if filename_draw:
            sub_map[filename_draw] = recs
        if internal and filename_draw and internal != filename_draw:
            drawing_alias_issues.append((os.path.basename(t['file']), filename_draw, internal))

    root_entries, top_entries, root_fallback = select_root_entries(all_entries, top_pdfs)
    if root_fallback:
        report.append('\n--- 根清单识别回退（需人工确认）---')
        report.append('  未找到未被引用的顶层PDF（可能存在循环/自引用），已合并全部可读顶层PDF，未丢弃任何一份。')
    elif not root_entries:
        # 理论上主PDF可读时不会进入；仍以全量而非首份兜底，避免静默遗漏。
        root_entries = list(all_entries)
        report.append('\n--- 根清单识别回退（需人工确认）---')
        report.append('  顶层PDF未能建立条目，已合并全部可读材料表，未丢弃任何一份。')
    if len(root_entries) > 1:
        report.append('\n--- 多根清单合并 ---')
        for entry in root_entries:
            report.append(f"  {entry['internal'] or entry['filename']} | {entry['table']['meta'].get('name', '')}")

    # 件号异常（缺失/格式不符）追加到复核报告：不重编号，只标记
    part_issues = []
    name_issues = []
    for entry in all_entries:
        d, recs = entry['internal'] or entry['filename'], entry['records']
        part_issues += [(d, r['name'], r.get('seq_raw', '')) for r in recs if not r['part_no']]
        name_issues += [(d, r.get('part_no') or r.get('seq_raw', '')) for r in recs if not r['name']]

    # ---- 子清单重量优先：独立材料表页脚总重覆盖总清单对应部件重量 ----
    sub_weight_map = {}
    sub_info_map = {}
    computed_totals = []
    detail_total_issues = []
    for entry in all_entries:
        sub_name, t, recs = entry['name'], entry['table'], entry['records']
        tw = t.get('total_weight')
        if tw is None:
            tw = extra_totals.get(sub_name)
        if tw is None:
            tw = detail_weight_total(recs)
            if tw is not None:
                computed_totals.append((entry['internal'] or entry['filename'], tw))
        else:
            detail_total = detail_weight_total(recs)
            if detail_total is not None and abs(detail_total - tw) > 0.005:
                detail_total_issues.append(
                    (entry['internal'] or entry['filename'], detail_total, tw))
        meta_qty = parse_int(str(t['meta'].get('qty', '')))
        info = {'weight': tw, 'meta_qty': meta_qty, 'entry': entry}
        for alias in {entry['internal'], entry['filename']} - {''}:
            sub_info_map[alias] = info
            if tw is not None:
                sub_weight_map[alias] = tw

    # 通过“父表总重 ≈ 子表单件总重 × 制造数量”校验并修复父件数量。
    qty_repairs = []
    def repair_parent_qty(records):
        for r in records:
            for ref in refs_of(r):
                refn = normalize_drawing(ref)
                info = sub_info_map.get(refn)
                if not info or not info['weight'] or r['weight'] is None:
                    continue
                unit_w = info['weight']
                ratio = r['weight'] / unit_w
                inferred = int(round(ratio))
                target = info['meta_qty'] if info['meta_qty'] else inferred
                tolerance = max(0.06, abs(unit_w * target) * 0.02)
                if target >= 1 and abs(r['weight'] - unit_w * target) <= tolerance and r['qty'] != target:
                    old = r['qty']
                    r['qty'] = target
                    qty_repairs.append((r['doc_drawing'], r['name'], old, target, refn))
                break
    for entry in all_entries:
        repair_parent_qty(entry['records'])

    overrides = []
    missing_totals = set()
    def override_weight(records):
        for r in records:
            for ref in refs_of(r):
                refn = normalize_drawing(ref)
                if refn in sub_weight_map:
                    r['weight_original'] = r['weight']
                    sub_unit = sub_weight_map[refn]
                    # 父表重量是该行全部数量的总重。与子表总重基本一致时，
                    # 用“父表总重÷数量”的完整精度作为单件重量，保证再乘后
                    # 精确回到父表原值；子表总重继续用于校验和缺失兜底。
                    main_unit = (r['weight'] / r['qty']) if r['weight'] is not None and r['qty'] else None
                    tolerance = max(0.06, abs(sub_unit) * 0.02)
                    if main_unit is not None and abs(main_unit - sub_unit) <= tolerance:
                        r['unit_weight'] = main_unit
                        r['weight_source'] = 'main_total_per_qty'
                    else:
                        r['unit_weight'] = sub_unit
                        r['weight_source'] = 'sub_pdf'
                    overrides.append((r['doc_drawing'], r['name'], r['weight_original'], r['unit_weight']))
                    break
                if refn in sub_map and refn not in missing_totals:
                    missing_totals.add(refn)   # 引用了子图但没捕获到子图总重
    for entry in all_entries:
        override_weight(entry['records'])

    # 把自动修正和仍需人工确认的内容统一写入报告。
    if drawing_alias_issues:
        report.append('\n--- 图号别名匹配（文件名与PDF内部材料表号不一致）---')
        for fn, alias, internal in drawing_alias_issues:
            report.append(f'  {fn} | 文件名={alias} | 内部材料表号={internal}（已按别名展开）')
    reread_changes = []
    seq_reread_changes = []
    for t in tables.values():
        for rec in t['records']:
            if '_qty_original' in rec:
                reread_changes.append((os.path.basename(t['file']), rec.get('name', ('', 0))[0],
                                       rec['_qty_original'], rec.get('qty', ('', 0))[0]))
            if '_seq_original' in rec:
                seq_reread_changes.append((os.path.basename(t['file']), rec.get('name', ('', 0))[0],
                                           rec['_seq_original'], rec.get('seq', ('', 0))[0]))
    if seq_reread_changes:
        report.append('\n--- 件号自动修正（高倍率复读）---')
        for fn, name, old, new in seq_reread_changes:
            report.append(f'  {fn} | {name} | {old} → {new}')
    if reread_changes or qty_repairs:
        report.append('\n--- 数量自动修正 ---')
        for fn, name, old, new in reread_changes:
            report.append(f'  {fn} | {name} | {old} → {new}（高倍率复读）')
        for d, name, old, new, ref in qty_repairs:
            report.append(f'  {d} | {name} | {old} → {new}（与子表 {ref} 总重/制造数量一致）')
    if computed_totals:
        report.append('\n--- 子清单总重回退计算（页脚未识别，按明细重量合计）---')
        for d, tw in computed_totals:
            report.append(f'  {d} | 合计={tw}')
    if detail_total_issues:
        report.append('\n--- PDF明细重量合计与页脚总重不一致（源表矛盾，未擅自改数）---')
        for d, detail_total, footer_total in detail_total_issues:
            report.append(f'  {d} | 明细合计={detail_total:.2f} | 页脚总重={footer_total:.2f}')
    if overrides:
        report.append('\n--- 装配件重量校验（父表总重保持精度，子表总重用于核对/兜底）---')
        for d, name, orig, new in overrides:
            report.append(f'  {d} | {name} | 总清单={orig} → 独立清单单件={new}')
    if missing_totals:
        report.append('\n--- 子清单总重未捕获（保持总清单重量，需人工核对）---')
        for d in sorted(missing_totals):
            report.append(f'  {d}')
    if part_issues:
        report.append('\n--- 件号复核（缺失/格式异常，代号只保留图号，需人工补件号）---')
        for d, name, raw in part_issues:
            report.append(f'  {d} | {name or "<名称空白>"} | OCR件号={raw!r}')
    if name_issues:
        report.append('\n--- 名称复核（PDF名称栏为空白或未识别，不自动猜测）---')
        for d, part in name_issues:
            report.append(f'  {d}-{part} | 名称为空白')
    with open(rp, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report))
    print(f'复核报告已保存: {rp}')

    # 层级序号严格使用 PDF 原件号，保留跳号和 2.1 等原始层级，不连续重编。
    def assign_numbers(parent):
        for c in parent.children:
            part = c.row.get('part_no') or '?'
            c.number = (parent.number + '.' if parent.number else '') + part
            assign_numbers(c)
    flat = []
    root_context = {}
    def flatten(node):
        if node.row:
            flat.append(node)
        for c in node.children:
            flatten(c)

    for entry in root_entries:
        root_draw = entry['internal'] or entry['filename']
        root_name = entry['table']['meta'].get('name', '')
        root = build_tree(entry['records'], sub_map, root_draw)
        assign_numbers(root)
        before = len(flat)
        flatten(root)
        for node in flat[before:]:
            root_context[id(node)] = (root_draw, root_name)

    # 同父件重名加序号（标准件除外）
    # 注意：同一子图被多个父件引用时，记录字典是共享的；用 id(row) 防重复处理，
    # 否则第二个父件会把 name_base 覆盖成带序号的名字（导致型材材质拼出假尺寸）
    children_of = defaultdict(list)
    for node in flat:
        children_of[id(node.parent)].append(node)
    numbered = set()
    for nodes in children_of.values():
        counts = Counter(n.row['name'] for n in nodes)
        seen = defaultdict(int)
        for node in nodes:
            rid = id(node.row)
            if rid in numbered:
                continue
            numbered.add(rid)
            nm = node.row['name']
            node.row['name_base'] = nm
            if nm and counts[nm] > 1 and not is_std_part(node.row):
                seen[nm] += 1
                node.row['name'] = f"{nm}{seen[nm]}"

    root_contracts = list(dict.fromkeys(
        entry['table']['meta'].get('contract', '') for entry in root_entries
        if entry['table']['meta'].get('contract', '')
    ))
    root_contract = '/'.join(root_contracts)
    if len(root_entries) == 1:
        output_root_name = root_entries[0]['table']['meta'].get('name', '')
        output_root_draw = root_entries[0]['internal'] or root_entries[0]['filename']
        title = f"{root_contract} {output_root_name}（{output_root_draw}）"
        out_name = f"工艺清单_{output_root_name}（{output_root_draw}）.xlsx"
    else:
        folder_name = os.path.basename(os.path.normpath(source_dir))
        title = f"{root_contract} {folder_name}（总清单）"
        out_name = f"工艺清单_{folder_name}（总清单）.xlsx"

    def parent_ref(node):
        p = node.parent
        if p is None or p.row is None:
            return root_context[id(node)]
        return (node_code(p), p.row['name'])

    def remark_of(r):
        return '' if r['ares'] else r['spec']

    # 生成 Excel
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb['模板表单']
    ws['C1'] = title

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    FONT = Font(name='宋体', size=12)
    FONT_B = Font(name='宋体', size=12, bold=True)
    ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

    r0 = 3
    row_of = {id(node): r0 + i for i, node in enumerate(flat)}
    for i, node in enumerate(flat):
        r = node.row
        xr = r0 + i
        ws.row_dimensions[xr].height = 26      # 数据行行高固定26磅（第1、2行保持模板）
        pref, pname = parent_ref(node)
        data = {
            1: node.number,
            2: node_code(node),
            3: r['name'],
            4: pref,
            5: pname,
            6: r['qty'],
            7: r['qty'] * node.mult,
            9: material_of(r),
            10: weight_per(r),
            14: remark_of(r),
        }
        bold = bool(node.children)
        for ci in range(1, 20):
            c = ws.cell(xr, ci)
            c.font = FONT_B if bold else FONT
            c.border = border
            c.alignment = ALIGN
            if ci in data:
                c.value = data[ci]
                if ci in (6, 7):
                    c.number_format = '0.###'
                elif ci in (10, 11):
                    c.number_format = '0.0000'
        # K 列（单台总重）：所有行统一 = 单台数量 × 单件重量（父件/子件/标准件一致）
        kc = ws.cell(xr, 11)
        kc.value = f'=G{xr}*J{xr}'
        kc.number_format = '0.0000'
    for xr in range(3, ws.max_row + 1):
        for ci in range(1, 20):
            c = ws.cell(xr, ci)
            if c.fill and c.fill.patternType:
                c.fill = PatternFill()

    out_path = os.path.join(source_dir, out_name)
    try:
        wb.save(out_path)
        print(f'\n完成! 合并根清单 {len(root_entries)} 个，共 {len(flat)} 行')
        print(f'输出: {out_path}')
    except PermissionError:
        alt = os.path.join(source_dir, out_name[:-5] + '（新）.xlsx')
        wb.save(alt)
        print(f'\n注意: 目标文件正被占用(可能WPS打开)，已另存为:\n  {alt}')
        print(f'完成! 共 {len(flat)} 行')

    # 打印前几行供核对
    for node in flat[:8]:
        r = node.row
        print(f"  {node.number:<6}{r['name']:<20} 数量{node.mult * r['qty']}  材质:{material_of(r)}")

if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else input('请输入图纸材料表PDF所在文件夹路径：').strip().strip('"')
    if not src:
        print('未输入路径')
        sys.exit(1)
    run(src)
