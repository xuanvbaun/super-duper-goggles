# -*- coding: utf-8 -*-
"""瑞士 Rowa 文本型 BOM PDF → 工艺清单 Excel
用法: python swiss_build.py "清单PDF所在文件夹"
规则(按参考文件V-26126-1681归纳):
  - L属性/M工艺内容/N备注 留空; A序号仅总装行; H计划数量仅总装行=1
  - ×/x 全部换成 *
  - 材质: BLE→钢板t{厚}/材质(S235JR+AR→Q235B, Hardox→NM400, Rubber→橡胶)
          RND→圆钢Ø{径}, ROR→圆管{外径}*{壁厚}, 标准件→GB/T标准-等级-ZC
          装配行→组件/组焊件
  - 代号: 图纸号'Pos (ZN子图→子图号'Pos); 标准件编码程序自定(用户后续统一改)
"""
import os, sys, re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(SCRIPT_DIR, '工艺清单标准模版-2026.6.30(1).xlsx')

TERMS = {
    'hopper panel welded': '料斗面板', 'hopper panel': '料斗面板', 'hopperpanel welded': '料斗面板', 'hopperpanel': '料斗面板',
    'hopper sidepanel welded': '料斗侧板', 'hopper sidepanel': '料斗侧板',
    'hopper cornerpanel': '料斗角板', 'cornerpanel': '角板',
    'hopper lid': '料斗盖板', 'lid': '盖板',
    'strip': '条板', 'fastening strip': '固定条', 'clamping strip': '夹紧条',
    'bearing plate': '轴承板', 'locking plate': '锁定板',
    'rib': '加强筋', 'chain holder': '链条托架', 'suspension eye': '吊环',
    'locking pin': '锁紧销', 'rubber rag': '橡胶垫块', 'grip': '握把',
    'round pipe welded': '圆管', 'round pipe': '圆管', 'ring': '圆环',
    'chute welded': '溜槽', 'chute panel': '溜槽面板', 'hardox plate': '耐磨板',
    'fastening bracket left': '左侧固定支架', 'fastening bracket right': '右侧固定支架',
    'hex head bolt': '螺栓', 'hex nut': '螺母', 'washer': '平垫', 'weld nut': '焊接螺母',
    'chain link': '链环',
}
MAT_MAP = {'S235JR+AR': 'Q235B', 'S235JR': 'Q235B', 'Hardox': 'NM400', 'Rubber': '橡胶', 'steel': '钢'}

def norm_name(n):
    n = n.lower().replace('_', ' ').replace('-', ' ').replace('.', ' ')
    n = re.sub(r'\s+', ' ', n).strip()
    n = n.replace('righgt', 'right')
    return n

def translate(name):
    n = norm_name(name)
    m = re.match(r'hopper (\d+) rvc sp10', n)
    if m:
        return f'料斗 {m.group(1)} RVC SP10'
    if n in TERMS:
        return TERMS[n]
    for k, v in TERMS.items():
        if n.startswith(k):
            rest = n[len(k):].strip().replace('x', '*').replace('×', '*')
            return v + rest
    return name

def std_part(pos_typ, length, werkstoff, remark, name):
    nm = norm_name(name)
    m = re.match(r'^M(\d+)', pos_typ)
    if not m:
        return None, None
    dia = m.group(1)
    grade = werkstoff.strip() if werkstoff.strip() in ('8.8', '8', '10.9', '12.9') else ('8.8' if 'bolt' in nm else '8')
    zc = 'ZC' if 'verzinkt' in remark else ''
    if 'hex head bolt' in nm:
        return f'螺栓M{dia}*{length}', f'GB/T5783-{grade}' + (f'-{zc}' if zc else '')
    if 'hex nut' in nm:
        return f'螺母M{dia}', f'GB/T6170-{grade}' + (f'-{zc}' if zc else '')
    if 'weld nut' in nm:
        return f'焊接螺母M{dia}', 'DIN-929'
    if 'washer' in nm:
        hv = re.search(r'(\d+)HV', werkstoff)
        hv = hv.group(1) if hv else '200'
        std = 'GB/T96.1' if ('7349' in remark or 'large' in nm) else 'GB/T97.1'
        name_cn = '大垫圈' if '7349' in remark or 'large' in nm else '平垫'
        return f'{name_cn}{dia}', f'{std}-{hv}HV' + (f'-{zc}' if zc else '')
    return None, None

def clean_len(s):
    if s is None: return ''
    s = str(s).strip()
    s = re.sub(r'^(x+|X+)', '', s)
    return s

def parse_pdf(path):
    pdf = pdfplumber.open(path)
    meta = {}
    t1 = pdf.pages[0].extract_text() or ''
    mz = re.search(r"Zeichnungsnummer:\s*([\d'．.]+)", t1)
    mt = re.search(r"Titel:\s*(.+)", t1)
    mp = re.search(r"Projekt:\s*(.+)", t1)
    mg = re.search(r"Gewicht total:\s*([\d.]+)", t1)
    meta['zeichnung'] = mz.group(1).strip().replace('．', "'") if mz else os.path.basename(path)
    meta['titel'] = mt.group(1).strip() if mt else ''
    meta['projekt'] = mp.group(1).strip() if mp else ''
    meta['gewicht_total'] = float(mg.group(1)) if mg else None

    rows = []
    for page in pdf.pages:
        for t in page.extract_tables() or []:
            for r in t:
                if not r: continue
                pos = str(r[0] or '').strip()
                if 'Pos' not in pos and not (len(r) > 5 and str(r[5] or '').strip().startswith('M')):
                    continue
                rows.append(r)
    return meta, rows

def g(r, i):
    if len(r) <= i or r[i] is None:
        return ''
    s = str(r[i]).strip()
    return '' if s == 'None' else s

def clean_row(r):
    pos = g(r, 0)
    m = re.search(r'(\d+)', pos)
    stk_m = re.search(r'\d+', g(r, 4))
    w_m = re.search(r'[\d.]+', g(r, 9))
    return {
        'pos': int(m.group(1)) if m else None,
        'name': g(r, 3), 'stk': int(stk_m.group()) if stk_m else 1,
        'typ': g(r, 5), 'dicke': g(r, 7), 'laenge': clean_len(g(r, 8)),
        'gewicht': float(w_m.group()) if w_m else None,
        'werkstoff': g(r, 11), 'bemerk': g(r, 12),
    }

def is_assembly(r, later):
    """组焊件: Pos<100 且其后存在 Pos>=100 的行（不依赖材质栏，材质可能为空或写了材质）"""
    if r['pos'] is None or r['pos'] >= 100:
        return False
    return any(p is not None and p >= 100 for p in later)

def build(pdf_path):
    meta, raw = parse_pdf(pdf_path)
    rows = [clean_row(r) for r in raw]
    later = [rows[i+1:] for i in range(len(rows))]

    root = {'name': translate(meta['titel']), 'code': meta['zeichnung'],
            'gewicht': meta['gewicht_total'], 'children': [], 'kind': 'root',
            'row': {'stk': 1, 'gewicht': meta['gewicht_total']}, 'parent': None}
    cur = root
    seq = 0
    for i, r in enumerate(rows):
        if is_assembly(r, [x['pos'] for x in later[i]]):
            # 组焊件是总装的同级子件（不是上一个组焊件的子件）
            node = {'name': translate(r['name']), 'gewicht': r['gewicht'], 'children': [],
                    'kind': 'weld', 'pos': r['pos'], 'row': r, 'parent': root}
            root['children'].append(node)
            cur = node
        else:
            if r['pos'] is not None and r['pos'] < 100:
                cur = root  # 非组焊件的 pos<100 行(直接件/ZN橡胶)回到总装，作为总装子件
            if r['pos'] is None:
                seq += 1
            node = {'name': translate(r['name']), 'gewicht': r['gewicht'], 'children': [],
                    'kind': 'part', 'pos': r['pos'], 'row': r, 'parent': cur, 'seq': seq}
            cur['children'].append(node)
    number_duplicates(root)
    return meta, root

def number_duplicates(node):
    """同一父件下同名的非标准件子件，名称后加序号区分"""
    from collections import defaultdict
    counts = defaultdict(int)
    for c in node['children']:
        cn, _ = std_part(c['row']['typ'], c['row']['laenge'], c['row']['werkstoff'], c['row']['bemerk'], c['row']['name'])
        if cn is None:
            counts[c['name']] += 1
    seen = defaultdict(int)
    for c in node['children']:
        cn, _ = std_part(c['row']['typ'], c['row']['laenge'], c['row']['werkstoff'], c['row']['bemerk'], c['row']['name'])
        if cn is None and counts[c['name']] > 1:
            seen[c['name']] += 1
            c['name'] = f"{c['name']}{seen[c['name']]}"
    for c in node['children']:
        number_duplicates(c)

def node_material(node):
    r = node['row']
    if node['kind'] == 'root':
        return '组件'
    if node['kind'] == 'weld':
        return '组焊件'
    typ = r['typ']
    w = MAT_MAP.get(r['werkstoff'], r['werkstoff'])
    cn, mat = std_part(typ, r['laenge'], r['werkstoff'], r['bemerk'], r['name'])
    if cn:
        return mat
    if typ.startswith('BLE'):
        if w == '橡胶':
            return '橡胶'
        d = r['dicke']
        return f"钢板t{d}/{w}" if (d and w) else (f"钢板t{d}" if d else (w or ''))
    m = re.match(r'RND(\d+)', typ)
    if m:
        return f"圆钢Ø{m.group(1)}/{w}" if w else f"圆钢Ø{m.group(1)}"
    m = re.match(r'ROR([\d.]+)x([\d.]+)', typ)
    if m:
        return f"圆管{m.group(1)}*{m.group(2)}/{w}" if w else f"圆管{m.group(1)}*{m.group(2)}"
    return w

def node_name(node):
    if node['kind'] == 'root':
        return node['name']
    r = node['row']
    cn, mat = std_part(r['typ'], r['laenge'], r['werkstoff'], r['bemerk'], r['name'])
    return cn if cn else node['name']

def node_code(node, meta):
    r = node['row']
    if node['kind'] == 'root':
        return meta['zeichnung']
    if node['kind'] == 'weld':
        return f"{meta['zeichnung']}'{r['pos']:03d}"
    if r['pos'] is None:
        return f"{meta['zeichnung']}'S{node.get('seq', 0):02d}"
    m = re.search(r"ZN\s*([\d']+)", r['bemerk'])
    if m:
        return f"{m.group(1).strip().replace('．', chr(39))}'{r['pos']:03d}"
    return f"{meta['zeichnung']}'{r['pos']:03d}"

def node_j(node):
    if node['kind'] == 'root':
        return node['row']['gewicht']
    w = node['row']['gewicht']
    if w is None:
        return None
    return round(w / node['row']['stk'], 6)

def flat_tree(root, out):
    out.append(root)
    for c in root['children']:
        flat_tree(c, out)

def build_excel(pdfs, out_path):
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb['模板表单']

    def sort_key(p):
        m = re.search(r"([\d']+)", os.path.basename(p))
        return m.group(1) if m else p
    all_nodes = []
    first_meta = None
    for pi, pdf_path in enumerate(sorted(pdfs, key=sort_key)):
        meta, root = build(pdf_path)
        if first_meta is None:
            first_meta = meta
        root['number'] = str(pi + 1)
        flat = []
        flat_tree(root, flat)
        for n in flat:
            n['_pdf'] = meta
        all_nodes.extend(flat)

    if first_meta:
        ws['C1'] = f"{first_meta['projekt']} {translate(first_meta['titel'])}（{first_meta['zeichnung']}）"

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    FONT = Font(name='宋体', size=10)
    FONT_B = Font(name='宋体', size=10, bold=True)
    ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

    idmap = {id(n): n for n in all_nodes}
    r0 = 3
    for i, node in enumerate(all_nodes):
        xr = r0 + i
        meta = node['_pdf']
        par = node.get('parent')
        pcode = pname = None
        if par is not None and par is not node:
            pcode = node_code(par, meta)
            pname = par['name']
        data = {
            1: node.get('number'),
            2: node_code(node, meta),
            3: node_name(node),
            4: pcode,
            5: pname,
            6: node['row']['stk'],
            7: node['row']['stk'],
            8: 1 if node['kind'] == 'root' else None,
            9: node_material(node),
            10: node_j(node),
            14: None,
        }
        bold = bool(node['children'])
        for ci in range(1, 20):
            c = ws.cell(xr, ci)
            c.font = FONT_B if bold else FONT
            c.border = border
            c.alignment = ALIGN
            if ci in data and data[ci] is not None:
                c.value = data[ci]
                if ci in (6, 7, 8):
                    c.number_format = '0'
        ws.cell(xr, 11).value = f'=G{xr}*J{xr}'

    for xr in range(3, ws.max_row + 1):
        for ci in range(1, 20):
            c = ws.cell(xr, ci)
            if c.fill and c.fill.patternType:
                c.fill = PatternFill()
    try:
        wb.save(out_path)
        print('已生成:', out_path)
    except PermissionError:
        alt = out_path[:-5] + '（新）.xlsx'
        wb.save(alt)
        print(f'注意: 目标文件正被占用(可能WPS打开)，已另存为:\n  {alt}')
    return out_path

def build_folder(src):
    """处理一个文件夹里的所有PDF(瑞士Rowa文本格式)，返回输出路径"""
    pdfs = [os.path.join(src, f) for f in os.listdir(src) if f.lower().endswith('.pdf')]
    if not pdfs:
        print('文件夹中没有PDF')
        return None
    meta0, _ = build(pdfs[0])
    out = os.path.join(src, f"工艺清单_{translate(meta0['titel'])}（{meta0['zeichnung']}）.xlsx")
    build_excel(pdfs, out)
    return out

if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else input('请输入清单PDF所在文件夹路径：').strip().strip('"')
    build_folder(src)
